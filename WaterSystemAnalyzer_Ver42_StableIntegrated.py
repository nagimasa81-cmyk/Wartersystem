# -*- coding: utf-8 -*-
"""
WaterSystem Analyzer Qt Ver40
=============================

Ver40 goals:
- Qt GUI, no tkinter.
- Worker-thread analysis; UI does not freeze.
- ResultSummary file selection is a file dialog.
- If a ResultSummary file is selected:
    * Analyze/update uses that file as the base.
    * Save updates that selected workbook path instead of creating a new file.
    * If serial number changes, the selected file is renamed and the selected-file path is updated.
- If no ResultSummary file is selected:
    * Analyze uses only the log files in the selected folder by default.
    * Existing ResultSummary is used only when the checkbox is enabled.
    * Save creates ResultSummary_<SN>.xlsx.
- CRC32 + file size duplicate prevention.
- X Start/End defines chart horizontal axis.
- Cum Start/End recalculates cumulative curves only.
- Checkboxes update chart immediately.
- Display Size presets.
"""

from __future__ import print_function
import uuid

import os
import zipfile
from pathlib import Path
import re
import sys
import glob
import zlib
import time
import shutil
import traceback
import json
import math
import datetime as dt

from PySide6 import QtCore, QtWidgets, QtGui
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QMessageBox, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QCheckBox, QComboBox, QScrollArea, QInputDialog,
    QProgressBar, QLineEdit, QGraphicsView, QGraphicsScene
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



APP_NAME = "WaterSystem Analyzer Version 7.3a DashboardDesignPreview"
APP_VERSION = "7.3a-DashboardDesignPreview"
ALGORITHM_VERSION = "Ver7.3a-DashboardDesignPreview"

RESULT_HEADERS = [
    "File Name",
    "Start Datetime (YYYY-MM-DD HH:MM)",
    "TREAT_CIRCULATE (HH:MM:SS)",
    "DEGAS_CIRCULATE (HH:MM:SS)",
    "CLEAN_TANK_CIRCULATE (HH:MM:SS)",
    "CLEAN_XD_CIRCULATE (HH:MM:SS)",
    "Total File Runtime (HH:MM:SS)",
    "Total File Runtime (Seconds)",
    "Total Circulate Time (Seconds)",
    "DEGAS_PAUSE (HH:MM:SS)",
    "TREAT_PAUSE (HH:MM:SS)",
    "ERROR (HH:MM:SS)",
    "Clean Count",
    "Primary Flow",
    "Chiller Temp",
    "Absolute Pressure",
    "Dynamic Pressure",
    "File CRC32",
    "File Size",
]


def normalize_space(s):
    return " ".join(str(s).replace("\t", " ").split())


def get_token(line, idx):
    p = normalize_space(line).split(" ")
    return p[idx].strip() if 0 <= idx < len(p) else ""


def parse_time_seconds(line):
    m = re.match(r"^\s*(\d{1,2}):(\d{2}):(\d{2})", str(line))
    if not m:
        return None
    return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))


def seconds_to_hms(sec):
    sec = int(round(float(sec or 0)))
    if sec < 0:
        sec = 0
    return "%02d:%02d:%02d" % (sec // 3600, (sec % 3600) // 60, sec % 60)


def hms_to_seconds(v):
    if isinstance(v, (int, float)):
        return float(v)
    m = re.match(r"^\s*(\d+):(\d{2}):(\d{2})", str(v or ""))
    if not m:
        return 0.0
    return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))


def month_value(s):
    return {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
    }.get(str(s).lower()[:3], 0)


def build_file_datetime_text(name):
    """
    Robust filename datetime parser.
    Supports weekday-included names:
      WaterSystem_Fri_Apr_10_11_44_23_2026.txt
      WaterSystem_Mon_Apr_20_18_48_50_2026.txt
      WaterSystem_Thu_Apr_02_08_28_03_2026.txt
    """
    base = os.path.splitext(os.path.basename(str(name)))[0]
    s = base.replace("-", "_").replace(".", "_").replace(" ", "_")
    tokens = [t for t in re.split(r"[_\W]+", s) if t]
    weekdays = set(["mon","monday","tue","tues","tuesday","wed","wednesday",
                    "thu","thur","thurs","thursday","fri","friday",
                    "sat","saturday","sun","sunday"])

    def make_dt(y, mo, d, hh=0, mm=0, ss=0):
        try:
            y = int(y); mo = int(mo); d = int(d)
            hh = int(hh); mm = int(mm); ss = int(ss)
            if y < 2000 or y > 2100:
                return None
            return dt.datetime(y, mo, d, hh, mm, ss)
        except Exception:
            return None

    # Token parse: optional prefix, optional weekday, month name, day, hh, mm, ss, year
    for i in range(len(tokens)):
        low = tokens[i].lower()

        # Month + day + hh + mm + ss + year
        mo = month_value(tokens[i])
        if mo and i + 5 < len(tokens):
            d, hh, mm, ss, y = tokens[i+1], tokens[i+2], tokens[i+3], tokens[i+4], tokens[i+5]
            if (re.fullmatch(r"\d{1,2}", str(d)) and re.fullmatch(r"\d{1,2}", str(hh)) and
                re.fullmatch(r"\d{1,2}", str(mm)) and re.fullmatch(r"\d{1,2}", str(ss)) and
                re.fullmatch(r"20\d{2}", str(y))):
                out = make_dt(y, mo, d, hh, mm, ss)
                if out:
                    return out

        # Weekday + Month + day + hh + mm + ss + year
        if low in weekdays and i + 6 < len(tokens):
            mo = month_value(tokens[i+1])
            if mo:
                d, hh, mm, ss, y = tokens[i+2], tokens[i+3], tokens[i+4], tokens[i+5], tokens[i+6]
                if (re.fullmatch(r"\d{1,2}", str(d)) and re.fullmatch(r"\d{1,2}", str(hh)) and
                    re.fullmatch(r"\d{1,2}", str(mm)) and re.fullmatch(r"\d{1,2}", str(ss)) and
                    re.fullmatch(r"20\d{2}", str(y))):
                    out = make_dt(y, mo, d, hh, mm, ss)
                    if out:
                        return out

    # Token parse: year + month text + day + hh + mm + ss
    # Example: 2026_Jul_05_10_22_57
    for i in range(len(tokens) - 5):
        if re.fullmatch(r"20\d{2}", str(tokens[i])):
            mo = month_value(tokens[i+1])
            if mo:
                d, hh, mm, ss = tokens[i+2], tokens[i+3], tokens[i+4], tokens[i+5]
                if (re.fullmatch(r"\d{1,2}", str(d)) and re.fullmatch(r"\d{1,2}", str(hh)) and
                    re.fullmatch(r"\d{1,2}", str(mm)) and re.fullmatch(r"\d{1,2}", str(ss))):
                    out = make_dt(tokens[i], mo, d, hh, mm, ss)
                    if out:
                        return out

    # Year first numeric: 2026_07_05_10_22_57
    m = re.search(r"(20\d{2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})(?:\D+(\d{1,2}))?", s)
    if m:
        out = make_dt(m.group(1), m.group(2), m.group(3), m.group(4), m.group(5), m.group(6) or 0)
        if out:
            return out

    # Year first month text: 2026_Jun_04_08_09_20
    m = re.search(r"(20\d{2})\D*([A-Za-z]{3,9})\D*(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})(?:\D+(\d{1,2}))?", s)
    if m:
        mo = month_value(m.group(2))
        if mo:
            out = make_dt(m.group(1), mo, m.group(3), m.group(4), m.group(5), m.group(6) or 0)
            if out:
                return out

    # Month text first, year last, no weekday
    m = re.search(r"([A-Za-z]{3,9})\D*(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(20\d{2})", s)
    if m:
        mo = month_value(m.group(1))
        if mo:
            out = make_dt(m.group(6), mo, m.group(2), m.group(3), m.group(4), m.group(5))
            if out:
                return out

    # Day-month text first, year last
    m = re.search(r"(\d{1,2})\D+([A-Za-z]{3,9})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(20\d{2})", s)
    if m:
        mo = month_value(m.group(2))
        if mo:
            out = make_dt(m.group(6), mo, m.group(1), m.group(3), m.group(4), m.group(5))
            if out:
                return out

    # Numeric month/day/time/year
    m = re.search(r"(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(20\d{2})", s)
    if m:
        out = make_dt(m.group(6), m.group(1), m.group(2), m.group(3), m.group(4), m.group(5))
        if out:
            return out

    return None


def coerce_datetime(value):
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in (
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%m/%d/%Y",
    ):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def is_watersystem_file(path):
    # FileReadFix: accept renamed WaterSystem logs.
    name = os.path.basename(path).lower()
    ext = os.path.splitext(name)[1].lower()
    if ext not in (".txt", ".log"):
        return False
    if "resultsummary" in name or name.startswith(("forecast", "replacement", "hospital")):
        return False
    return True


def scan_folder_diagnostics(folder):
    total = txtlog = accepted = resultsummary = 0
    samples = []
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if d.lower() not in ("$recycle.bin", "system volume information", "__pycache__")]
        for name in files:
            total += 1
            p = os.path.join(root, name)
            if os.path.splitext(name)[1].lower() in (".txt", ".log"):
                txtlog += 1
            if "resultsummary" in name.lower() and os.path.splitext(name)[1].lower() in (".xlsx", ".xlsm"):
                resultsummary += 1
            if is_watersystem_file(p):
                accepted += 1
                if len(samples) < 10:
                    samples.append(p)
    return {"total": total, "txtlog": txtlog, "accepted": accepted, "resultsummary": resultsummary, "samples": samples}


def file_crc32_hex(path):
    crc = 0
    with open(path, "rb") as f:
        for b in iter(lambda: f.read(1024 * 1024), b""):
            crc = zlib.crc32(b, crc)
    return "%08X" % (crc & 0xFFFFFFFF)


def content_key(h, size):
    if not h or size in ("", None):
        return ""
    return "CRC32SIZE|%s|%s" % (str(h).upper(), str(size))


def result_key(file_name, file_dt):
    d = file_dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(file_dt, dt.datetime) else str(file_dt or "")
    return "FILEDATE|%s|%s" % (str(file_name).lower().strip(), d.lower().strip())


def name_key(file_name):
    return "NAME|%s" % str(file_name).lower().strip()


def safe_sn(sn):
    sn = str(sn or "NA").strip() or "NA"
    return re.sub(r'[\\/:*?"<>|]+', "_", sn)


def resultsummary_mode_text(explicit_result_file, use_existing_result):
    if explicit_result_file:
        return "Explicit selected ResultSummary"
    if use_existing_result:
        return "Latest ResultSummary in selected folder"
    return "Log files only / no existing ResultSummary"


def resolve_resultsummary_base_file(folder, explicit_result_file, use_existing_result):
    """
    Step4 ResultSummary base-file policy.

    Priority:
    1. Explicit selected ResultSummary
    2. Latest ResultSummary in folder only when checkbox is ON
    3. Empty string: analyze log files only
    """
    if explicit_result_file:
        return explicit_result_file
    if use_existing_result:
        return find_latest_result_summary(folder) or ""
    return ""


def resultsummary_mode_text(explicit_result_file, use_existing_result):
    if explicit_result_file:
        return "Explicit selected ResultSummary"
    if use_existing_result:
        return "Latest ResultSummary in selected folder"
    return "Log files only / no existing ResultSummary"


def result_summary_path_for_sn(folder, sn):
    return os.path.join(folder, "ResultSummary_%s.xlsx" % safe_sn(sn))



def write_startup_log(message):
    try:
        base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base, "startup_error.log")
        with open(path, "a", encoding="utf-8") as f:
            f.write("[%s] %s\n" % (dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S"), message))
    except Exception:
        pass


def app_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))



def app_settings_path():
    return os.path.join(app_base_dir(), "WaterSystemAnalyzer_Settings.ini")


def default_data_folder():
    return os.path.join(app_base_dir(), "Data")


def get_data_root():
    """
    Step6.8:
    Default Data folder is under the EXE/application folder.
    User can move Data folder. The selected location is stored in WaterSystemAnalyzer_Settings.ini.
    """
    default_root = default_data_folder()
    path = default_root
    try:
        settings = app_settings_path()
        if os.path.exists(settings):
            with open(settings, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip().startswith("DataFolder="):
                        candidate = line.split("=", 1)[1].strip()
                        if candidate:
                            path = candidate
                        break
    except Exception:
        path = default_root
    return path or default_root


def set_data_root(path):
    path = os.path.abspath(path)
    os.makedirs(path, exist_ok=True)
    with open(app_settings_path(), "w", encoding="utf-8") as f:
        f.write("DataFolder=%s\n" % path)
        f.write("APP_VERSION=%s\n" % APP_VERSION)
        f.write("UPDATED=%s\n" % dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S"))
    return path



def master_foundation_path(name):
    root = ensure_data_folders()
    master = os.path.join(root, "Master")
    os.makedirs(master, exist_ok=True)
    return os.path.join(master, name)


HP_MASTER_HEADERS = ["Hospital Name", "Site", "Serial Number", "Installation Date", "Coil Type", "Status", "Comment"]
REPLACEMENT_HISTORY_HEADERS = ["Hospital Name", "Serial Number", "Component Type", "Component ID", "Coil ID", "Replacement Date", "Replacement Type", "Reason", "Comment"]
COMPONENT_MASTER_HEADERS = ["Hospital Name", "Serial Number", "Component Type", "Component ID", "Install Date", "Status", "Comment"]



def migrate_legacy_hp_master():
    """Version 7.0.1: migrate legacy HP_Master.xlsx to Hospital_Master.xlsx if needed."""
    try:
        root = ensure_data_folders()
        master = os.path.join(root, "Master")
        old_path = os.path.join(master, "HP_Master.xlsx")
        new_path = os.path.join(master, "Hospital_Master.xlsx")
        if os.path.exists(old_path) and not os.path.exists(new_path):
            shutil.copy2(old_path, new_path)
    except Exception:
        pass


def ensure_master_foundation_files():
    migrate_legacy_hp_master()
    """
    Step5 only creates and maintains master input files.
    It does not calculate Forecast or MTBF.
    """
    specs = {
        "Hospital_Master.xlsx": HP_MASTER_HEADERS,
        "Component_Master.xlsx": COMPONENT_MASTER_HEADERS,
        "Replacement_History.xlsx": REPLACEMENT_HISTORY_HEADERS,
    }
    for fname, headers in specs.items():
        path = master_foundation_path(fname)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(fname)[0]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            ws.freeze_panes = "A2"
            for c in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(headers[c - 1]) + 3)
            wb.save(path)
        else:
            ensure_headers(path, headers)


def read_master_foundation_rows(filename, headers):
    ensure_master_foundation_files()
    path = master_foundation_path(filename)
    rows = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        sheet_headers = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            if not any(ws.cell(r, c).value not in ("", None) for c in range(1, ws.max_column + 1)):
                continue
            row = {}
            for c, h in enumerate(sheet_headers, 1):
                if h:
                    row[h] = ws.cell(r, c).value
            rows.append(row)
    except Exception:
        pass
    return rows


def write_master_foundation_rows(filename, headers, rows):
    ensure_master_foundation_files()
    path = master_foundation_path(filename)
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(filename)[0]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for r_idx, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(r_idx, c, row.get(h, ""))
    ws.freeze_panes = "A2"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(headers[c - 1]) + 3)
    wb.save(path)
    return path


def excel_serial_to_datetime(value):
    """
    Convert Excel serial date or common date text to datetime.
    History.xlsx uses Excel date serial values in date columns.
    """
    if value in ("", None):
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    try:
        # Excel serial date system. 25569 = 1970-01-01.
        if isinstance(value, (int, float)) and value > 20000:
            return dt.datetime(1899, 12, 30) + dt.timedelta(days=float(value))
    except Exception:
        pass

    s = str(value).strip()
    if not s:
        return None
    # Guard impossible typo-like strings such as 2025/07/61.
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        if re.fullmatch(r"\d+(\.\d+)?", s):
            num = float(s)
            if num > 20000:
                return dt.datetime(1899, 12, 30) + dt.timedelta(days=num)
    except Exception:
        pass
    return None


def normalize_history_key(value):
    if value in ("", None):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()



def calculate_data_coverage_skeleton(summary_rows):
    dates = [r.get("DateTime") for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)]
    if not dates:
        return {"First Log Date": "", "Last Log Date": "", "Coverage Days": 0, "Estimated Usage Candidate": "No log dates"}
    first = min(dates)
    last = max(dates)
    return {"First Log Date": first.strftime("%Y/%m/%d"), "Last Log Date": last.strftime("%Y/%m/%d"), "Coverage Days": (last-first).days + 1, "Estimated Usage Candidate": "Prepared for Step7"}




def normalize_text_value(v):
    if v is None:
        return ""
    try:
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
    except Exception:
        pass
    return str(v).strip()


def parse_date_value(v):
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    if v in ("", None):
        return None
    s = str(v).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return coerce_datetime(s)
    except Exception:
        return None


# ==============================
# Version 7.1 Reliability Engine
# ==============================

REPLACEMENT_VALIDATION_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Replacement Date", "Replacement Type", "Status",
    "Duplicate Check", "Date Check", "Mapping Check", "Warning", "Source"
]

GAP_ANALYSIS_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Gap Start", "Gap End", "Gap Days", "Handling",
    "Estimated Runtime Hours", "Basis", "Confidence Impact", "Warning"
]

RELIABILITY_HISTORY_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Coverage %", "Gap Score", "Replacement Score", "Runtime Continuity Score",
    "Log Quality Score", "Confidence Score", "Confidence Stars",
    "Data Quality", "Estimated Runtime Hours", "Warning", "Algorithm Version"
]

FORECAST_SNAPSHOT_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Observed Runtime Hours", "Estimated Runtime Hours", "Total Runtime Basis Hours",
    "Coverage %", "Confidence Score", "Confidence Stars",
    "Forecast Basis", "Explanation", "Algorithm Version"
]


def ensure_reliability_files():
    specs = {
        "Replacement_Validation.xlsx": REPLACEMENT_VALIDATION_HEADERS,
        "Gap_Analysis.xlsx": GAP_ANALYSIS_HEADERS,
        "Reliability_History.xlsx": RELIABILITY_HISTORY_HEADERS,
        "Forecast_Snapshot.xlsx": FORECAST_SNAPSHOT_HEADERS,
    }
    for fname, headers in specs.items():
        path = master_path(fname)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(fname)[0]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 3)
            ws.freeze_panes = "A2"
            wb.save(path)


def append_reliability_rows(fname, headers, rows):
    ensure_reliability_files()
    path = master_path(fname)
    wb = load_workbook(path)
    ws = wb.active
    for row in rows:
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    wb.save(path)
    return path


def confidence_stars(score):
    try:
        score = float(score)
    except Exception:
        score = 0
    if score >= 95:
        return "★★★★★"
    if score >= 85:
        return "★★★★☆"
    if score >= 70:
        return "★★★☆☆"
    if score >= 50:
        return "★★☆☆☆"
    return "★☆☆☆☆"


def read_replacement_history_rows():
    # Primary V7 file name; read via current header layout.
    try:
        return read_master_foundation_rows("Replacement_History.xlsx", REPLACEMENT_HISTORY_HEADERS)
    except Exception:
        return []


def validate_replacement_history():
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    rows = read_replacement_history_rows()
    seen = set()
    out = []
    for r in rows:
        hospital = normalize_text_value(r.get("Hospital Name") or r.get("Hospital") or r.get("Site"))
        sn = normalize_text_value(r.get("Serial Number") or r.get("SN"))
        ctype = normalize_text_value(r.get("Component Type") or r.get("Part")) or "Degas Module"
        rdate = parse_date_value(r.get("Replacement Date"))
        rtype = normalize_text_value(r.get("Replacement Type")) or "Unknown"

        key = (hospital, sn, ctype, rdate.strftime("%Y/%m/%d") if rdate else "")
        duplicate = "Duplicate" if key in seen and key[-1] else "OK"
        if key[-1]:
            seen.add(key)

        warnings = []
        date_check = "OK" if rdate else "Missing/Invalid"
        mapping_check = "OK" if (hospital and sn and ctype) else "Missing mapping"
        if date_check != "OK":
            warnings.append("Replacement date is missing or invalid.")
        if mapping_check != "OK":
            warnings.append("Hospital Name / Serial Number / Component Type is incomplete.")
        if duplicate != "OK":
            warnings.append("Possible duplicate replacement record.")

        status = "OK" if not warnings else "Warning"
        out.append({
            "Run Date": run_date,
            "Hospital Name": hospital,
            "Serial Number": sn,
            "Component Type": ctype,
            "Replacement Date": rdate.strftime("%Y/%m/%d") if rdate else "",
            "Replacement Type": rtype,
            "Status": status,
            "Duplicate Check": duplicate,
            "Date Check": date_check,
            "Mapping Check": mapping_check,
            "Warning": " ".join(warnings),
            "Source": "Replacement_History.xlsx",
        })
    return out


def detect_log_gaps(summary_rows, gap_threshold_days=3, missing_handling="Unknown"):
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    dates = sorted({r.get("DateTime").date() for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)})
    if len(dates) < 2:
        return []
    gaps = []
    for a, b in zip(dates, dates[1:]):
        missing_days = (b - a).days - 1
        if missing_days >= gap_threshold_days:
            gap_start = a + dt.timedelta(days=1)
            gap_end = b - dt.timedelta(days=1)
            if missing_handling == "Estimated Usage":
                # Conservative initial estimate: use average Runtime daily value when available.
                runtime_vals = [float(r.get("Runtime", 0) or 0) for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)]
                avg_runtime = sum(runtime_vals) / len(runtime_vals) if runtime_vals else 0.0
                est_hours = round(avg_runtime * missing_days, 2)
                basis = "Average observed daily Runtime x missing days"
                impact = "Medium"
            elif missing_handling == "Conservative":
                runtime_vals = [float(r.get("Runtime", 0) or 0) for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)]
                avg_runtime = sum(runtime_vals) / len(runtime_vals) if runtime_vals else 0.0
                est_hours = round(avg_runtime * missing_days * 0.5, 2)
                basis = "50% of average observed daily Runtime x missing days"
                impact = "Medium"
            else:
                est_hours = 0.0
                basis = "Unknown gap / no usage estimate"
                impact = "High"

            gaps.append({
                "Run Date": run_date,
                "Hospital Name": "",
                "Serial Number": "",
                "Component Type": "Degas Module",
                "Gap Start": gap_start.strftime("%Y/%m/%d"),
                "Gap End": gap_end.strftime("%Y/%m/%d"),
                "Gap Days": missing_days,
                "Handling": missing_handling,
                "Estimated Runtime Hours": est_hours,
                "Basis": basis,
                "Confidence Impact": impact,
                "Warning": "Log gap detected.",
            })
    return gaps


def build_reliability_dashboard(summary_rows, missing_handling="Unknown"):
    ensure_reliability_files()
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    dates = sorted({r.get("DateTime").date() for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)})
    if dates:
        first = dates[0]
        last = dates[-1]
        total_days = (last - first).days + 1
        observed_days = len(dates)
        coverage = round((observed_days / max(1, total_days)) * 100.0, 1)
    else:
        first = last = None
        total_days = observed_days = 0
        coverage = 0.0

    validation_rows = validate_replacement_history()
    warning_count = sum(1 for r in validation_rows if r.get("Status") != "OK")
    replacement_score = max(0, 100 - warning_count * 15)
    gaps = detect_log_gaps(summary_rows, 3, missing_handling)
    gap_days = sum(int(g.get("Gap Days", 0) or 0) for g in gaps)
    estimated_hours = round(sum(float(g.get("Estimated Runtime Hours", 0) or 0) for g in gaps), 2)

    coverage_score = min(100, coverage)
    gap_score = max(0, 100 - gap_days * 2)
    runtime_score = 100 if observed_days >= 30 else (70 if observed_days >= 7 else 40)
    log_quality_score = 100 if coverage >= 85 else (75 if coverage >= 60 else 45)

    confidence_score = round(
        coverage_score * 0.25 +
        gap_score * 0.25 +
        replacement_score * 0.20 +
        runtime_score * 0.15 +
        log_quality_score * 0.15,
        1
    )

    stars = confidence_stars(confidence_score)
    data_quality = "Excellent" if confidence_score >= 90 else ("Good" if confidence_score >= 75 else ("Fair" if confidence_score >= 55 else "Poor"))
    warnings = []
    if warning_count:
        warnings.append("%d replacement warning(s)" % warning_count)
    if gap_days:
        warnings.append("%d missing log day(s)" % gap_days)
    if not dates:
        warnings.append("No log data loaded")

    reliability_row = {
        "Run Date": run_date,
        "Hospital Name": "",
        "Serial Number": "",
        "Component Type": "Degas Module",
        "Coverage %": coverage,
        "Gap Score": gap_score,
        "Replacement Score": replacement_score,
        "Runtime Continuity Score": runtime_score,
        "Log Quality Score": log_quality_score,
        "Confidence Score": confidence_score,
        "Confidence Stars": stars,
        "Data Quality": data_quality,
        "Estimated Runtime Hours": estimated_hours,
        "Warning": " / ".join(warnings),
        "Algorithm Version": ALGORITHM_VERSION,
    }

    observed_runtime = round(sum(float(r.get("Runtime", 0) or 0) for r in (summary_rows or [])), 2)
    snapshot_row = {
        "Run Date": run_date,
        "Hospital Name": "",
        "Serial Number": "",
        "Component Type": "Degas Module",
        "Observed Runtime Hours": observed_runtime,
        "Estimated Runtime Hours": estimated_hours,
        "Total Runtime Basis Hours": round(observed_runtime + estimated_hours, 2),
        "Coverage %": coverage,
        "Confidence Score": confidence_score,
        "Confidence Stars": stars,
        "Forecast Basis": "Measured + Estimated" if estimated_hours else "Measured only",
        "Explanation": "Observed runtime %.2f h + estimated gap %.2f h. Coverage %.1f%%. Confidence %s." % (observed_runtime, estimated_hours, coverage, stars),
        "Algorithm Version": ALGORITHM_VERSION,
    }

    return validation_rows, gaps, [reliability_row], [snapshot_row]


def save_reliability_dashboard(summary_rows, missing_handling="Unknown"):
    validation_rows, gap_rows, reliability_rows, snapshot_rows = build_reliability_dashboard(summary_rows, missing_handling)
    p1 = append_reliability_rows("Replacement_Validation.xlsx", REPLACEMENT_VALIDATION_HEADERS, validation_rows)
    p2 = append_reliability_rows("Gap_Analysis.xlsx", GAP_ANALYSIS_HEADERS, gap_rows)
    p3 = append_reliability_rows("Reliability_History.xlsx", RELIABILITY_HISTORY_HEADERS, reliability_rows)
    p4 = append_reliability_rows("Forecast_Snapshot.xlsx", FORECAST_SNAPSHOT_HEADERS, snapshot_rows)
    return (p1, p2, p3, p4), validation_rows, gap_rows, reliability_rows, snapshot_rows



RUNTIME_RECONSTRUCTION_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Period Start", "Period End", "Runtime Type",
    "Measured Runtime Hours", "Estimated Runtime Hours", "Confirmed Stop Hours",
    "Unknown Days", "Total Runtime Basis Hours",
    "Handling", "Basis", "Confidence Impact", "Warning", "Algorithm Version"
]
GAP_HISTORY_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Gap Start", "Gap End", "Gap Days", "Handling",
    "Estimated Runtime Hours", "Basis", "Resolved By", "Algorithm Version"
]
PREDICTION_HISTORY_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Observed Runtime Hours", "Estimated Runtime Hours", "Total Runtime Basis Hours",
    "Forecast Method", "Prediction Status", "Prediction Note", "Algorithm Version"
]
CONFIDENCE_HISTORY_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Coverage Score", "Gap Quality Score", "Replacement History Score",
    "Runtime Quality Score", "Prediction Stability Score",
    "Overall Confidence Score", "Overall Confidence Stars",
    "Explanation", "Algorithm Version"
]

def ensure_runtime_reconstruction_files():
    specs = {
        "Runtime_Reconstruction.xlsx": RUNTIME_RECONSTRUCTION_HEADERS,
        "Gap_History.xlsx": GAP_HISTORY_HEADERS,
        "Prediction_History.xlsx": PREDICTION_HISTORY_HEADERS,
        "Confidence_History.xlsx": CONFIDENCE_HISTORY_HEADERS,
    }
    for fname, headers in specs.items():
        path = master_path(fname)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(fname)[0]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 3)
            ws.freeze_panes = "A2"
            wb.save(path)

def append_runtime_rows(fname, headers, rows):
    ensure_runtime_reconstruction_files()
    path = master_path(fname)
    wb = load_workbook(path)
    ws = wb.active
    for row in rows:
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    wb.save(path)
    return path

def runtime_daily_map(summary_rows):
    daily = {}
    for r in summary_rows or []:
        d = r.get("DateTime")
        if not isinstance(d, dt.datetime):
            continue
        day = d.date()
        try:
            val = float(r.get("Runtime", 0) or 0)
        except Exception:
            val = 0.0
        if val == 0:
            for key in ["Treat", "Degas", "Clean"]:
                try:
                    val += float(r.get(key, 0) or 0)
                except Exception:
                    pass
        daily[day] = daily.get(day, 0.0) + val
    return daily

def build_runtime_reconstruction(summary_rows, missing_handling="Unknown"):
    ensure_runtime_reconstruction_files()
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    daily = runtime_daily_map(summary_rows)
    days = sorted(daily.keys())
    runtime_rows = []
    gap_rows = []

    if not days:
        runtime_rows.append({
            "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
            "Component Type": "Degas Module", "Period Start": "", "Period End": "",
            "Runtime Type": "Unknown Runtime", "Measured Runtime Hours": 0,
            "Estimated Runtime Hours": 0, "Confirmed Stop Hours": 0,
            "Unknown Days": 0, "Total Runtime Basis Hours": 0,
            "Handling": missing_handling, "Basis": "No log data",
            "Confidence Impact": "High", "Warning": "No runtime data available.",
            "Algorithm Version": ALGORITHM_VERSION,
        })
        return runtime_rows, gap_rows

    measured_total = round(sum(daily.values()), 2)
    runtime_rows.append({
        "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
        "Component Type": "Degas Module", "Period Start": days[0].strftime("%Y/%m/%d"),
        "Period End": days[-1].strftime("%Y/%m/%d"), "Runtime Type": "Measured Runtime",
        "Measured Runtime Hours": measured_total, "Estimated Runtime Hours": 0,
        "Confirmed Stop Hours": 0, "Unknown Days": 0,
        "Total Runtime Basis Hours": measured_total, "Handling": "Measured",
        "Basis": "Runtime from ResultSummary/log data", "Confidence Impact": "Low",
        "Warning": "", "Algorithm Version": ALGORITHM_VERSION,
    })

    avg_runtime = measured_total / max(1, len(days))
    for a, b in zip(days, days[1:]):
        missing_days = (b - a).days - 1
        if missing_days <= 0:
            continue
        gap_start = a + dt.timedelta(days=1)
        gap_end = b - dt.timedelta(days=1)
        if missing_handling == "Estimated Usage":
            est = round(avg_runtime * missing_days, 2)
            rt_type = "Estimated Runtime"
            basis = "Average measured daily runtime x gap days"
            impact = "Medium"
            unknown_days = 0
        elif missing_handling == "Conservative":
            est = round(avg_runtime * missing_days * 0.5, 2)
            rt_type = "Estimated Runtime"
            basis = "50% average measured daily runtime x gap days"
            impact = "Medium"
            unknown_days = 0
        else:
            est = 0.0
            rt_type = "Unknown Runtime"
            basis = "Gap is treated as unknown; no runtime estimate"
            impact = "High"
            unknown_days = missing_days

        runtime_rows.append({
            "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
            "Component Type": "Degas Module", "Period Start": gap_start.strftime("%Y/%m/%d"),
            "Period End": gap_end.strftime("%Y/%m/%d"), "Runtime Type": rt_type,
            "Measured Runtime Hours": 0, "Estimated Runtime Hours": est,
            "Confirmed Stop Hours": 0, "Unknown Days": unknown_days,
            "Total Runtime Basis Hours": est, "Handling": missing_handling,
            "Basis": basis, "Confidence Impact": impact,
            "Warning": "Log gap reconstructed.", "Algorithm Version": ALGORITHM_VERSION,
        })
        gap_rows.append({
            "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
            "Component Type": "Degas Module", "Gap Start": gap_start.strftime("%Y/%m/%d"),
            "Gap End": gap_end.strftime("%Y/%m/%d"), "Gap Days": missing_days,
            "Handling": missing_handling, "Estimated Runtime Hours": est,
            "Basis": basis, "Resolved By": "Runtime Reconstruction Engine",
            "Algorithm Version": ALGORITHM_VERSION,
        })
    return runtime_rows, gap_rows


# =====================================
# Version 7.2c Explainable Reliability Engine
# =====================================

EXPLAINABLE_RELIABILITY_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Forecast Readiness", "Forecast Readiness Score",
    "Overall Confidence Score", "Overall Confidence Stars",
    "Coverage Score", "Coverage Reason",
    "Data Continuity Score", "Data Continuity Reason",
    "Gap Quality Score", "Gap Quality Reason",
    "Runtime Confidence Score", "Runtime Confidence Reason",
    "Replacement Quality Score", "Replacement Quality Reason",
    "Prediction Stability Score", "Prediction Stability Reason",
    "Data Quality Score", "Data Quality Reason",
    "Overall Explanation", "Algorithm Version"
]

RELIABILITY_TREND_HEADERS = [
    "Run Date", "Hospital Name", "Serial Number", "Component Type",
    "Forecast Readiness", "Overall Confidence Score", "Overall Confidence Stars",
    "Coverage Score", "Data Continuity Score", "Gap Quality Score",
    "Runtime Confidence Score", "Replacement Quality Score",
    "Prediction Stability Score", "Data Quality Score",
    "Measured Runtime Hours", "Estimated Runtime Hours", "Unknown Days",
    "Algorithm Version"
]


def ensure_explainable_reliability_files():
    specs = {
        "Explainable_Reliability.xlsx": EXPLAINABLE_RELIABILITY_HEADERS,
        "Reliability_Trend.xlsx": RELIABILITY_TREND_HEADERS,
    }
    for fname, headers in specs.items():
        path = master_path(fname)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(fname)[0]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 3)
            ws.freeze_panes = "A2"
            wb.save(path)


def append_explainable_rows(fname, headers, rows):
    ensure_explainable_reliability_files()
    path = master_path(fname)
    wb = load_workbook(path)
    ws = wb.active
    for row in rows:
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    wb.save(path)
    return path


def score_readiness(score, unknown_days, replacement_score):
    try:
        score = float(score)
    except Exception:
        score = 0
    try:
        unknown_days = int(unknown_days or 0)
    except Exception:
        unknown_days = 0
    try:
        replacement_score = float(replacement_score)
    except Exception:
        replacement_score = 0

    if score >= 85 and unknown_days <= 7 and replacement_score >= 80:
        return "READY"
    if score >= 65 and unknown_days <= 45:
        return "READY WITH CAUTION"
    return "NOT READY"


def build_explainable_reliability(summary_rows, runtime_rows, gap_rows, replacement_validation_rows=None, prediction_rows=None):
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    replacement_validation_rows = replacement_validation_rows or []
    prediction_rows = prediction_rows or []

    daily = runtime_daily_map(summary_rows)
    days = sorted(daily.keys())
    if days:
        total_span_days = (days[-1] - days[0]).days + 1
        observed_days = len(days)
        coverage_score = round((observed_days / max(1, total_span_days)) * 100.0, 1)
        coverage_reason = "%d observed day(s) over %d calendar day(s)." % (observed_days, total_span_days)
    else:
        total_span_days = 0
        observed_days = 0
        coverage_score = 0.0
        coverage_reason = "No log date data."

    gap_count = len(gap_rows or [])
    gap_days = sum(int(g.get("Gap Days", 0) or 0) for g in (gap_rows or []))
    max_gap = max([int(g.get("Gap Days", 0) or 0) for g in (gap_rows or [])] + [0])

    data_continuity_score = max(0, 100 - gap_count * 8 - max_gap * 1.5)
    data_continuity_reason = "%d gap(s), maximum gap %d day(s)." % (gap_count, max_gap)

    gap_quality_score = max(0, 100 - gap_days * 2)
    gap_quality_reason = "%d total missing day(s)." % gap_days

    measured = round(sum(float(r.get("Measured Runtime Hours", 0) or 0) for r in (runtime_rows or [])), 2)
    estimated = round(sum(float(r.get("Estimated Runtime Hours", 0) or 0) for r in (runtime_rows or [])), 2)
    unknown_days = sum(int(r.get("Unknown Days", 0) or 0) for r in (runtime_rows or []))
    total_runtime = measured + estimated

    if total_runtime > 0:
        measured_ratio = measured / total_runtime
        runtime_confidence_score = round(max(0, min(100, measured_ratio * 100 - unknown_days * 1.5)), 1)
        runtime_reason = "Measured runtime %.1f%% of runtime basis; unknown days %d." % (measured_ratio * 100, unknown_days)
    else:
        runtime_confidence_score = 0.0
        runtime_reason = "No runtime basis data."

    rep_rows = replacement_validation_rows
    rep_warning_count = sum(1 for r in rep_rows if r.get("Status") != "OK")
    if rep_rows:
        replacement_quality_score = max(0, 100 - rep_warning_count * 18)
        replacement_reason = "%d replacement record(s), %d warning(s)." % (len(rep_rows), rep_warning_count)
    else:
        replacement_quality_score = 50
        replacement_reason = "No replacement validation rows; neutral initial score."

    # Until prediction dates are available, use neutral stability score.
    prediction_stability_score = 50
    prediction_stability_reason = "Prediction date is not calculated yet; neutral initial score."

    # Basic data quality checks: negative runtime, duplicate days not detectable after aggregation but negative values are.
    negative_count = 0
    for r in summary_rows or []:
        for key in ["Runtime", "Treat", "Degas", "Clean"]:
            try:
                if float(r.get(key, 0) or 0) < 0:
                    negative_count += 1
            except Exception:
                pass
    data_quality_score = max(0, 100 - negative_count * 10)
    data_quality_reason = "%d negative runtime value(s) detected." % negative_count if negative_count else "No basic runtime anomaly detected."

    overall = round(
        coverage_score * 0.15 +
        data_continuity_score * 0.15 +
        gap_quality_score * 0.15 +
        replacement_quality_score * 0.20 +
        runtime_confidence_score * 0.20 +
        prediction_stability_score * 0.15,
        1
    )
    stars = confidence_stars(overall)
    readiness = score_readiness(overall, unknown_days, replacement_quality_score)

    explanation = (
        "Overall %.1f (%s). Coverage %.1f because %s "
        "Continuity %.1f because %s "
        "Gap %.1f because %s "
        "Runtime %.1f because %s "
        "Replacement %.1f because %s "
        "Prediction stability %.1f because %s"
    ) % (
        overall, stars,
        coverage_score, coverage_reason,
        data_continuity_score, data_continuity_reason,
        gap_quality_score, gap_quality_reason,
        runtime_confidence_score, runtime_reason,
        replacement_quality_score, replacement_reason,
        prediction_stability_score, prediction_stability_reason
    )

    row = {
        "Run Date": run_date,
        "Hospital Name": "",
        "Serial Number": "",
        "Component Type": "Degas Module",
        "Forecast Readiness": readiness,
        "Forecast Readiness Score": overall,
        "Overall Confidence Score": overall,
        "Overall Confidence Stars": stars,
        "Coverage Score": coverage_score,
        "Coverage Reason": coverage_reason,
        "Data Continuity Score": round(data_continuity_score, 1),
        "Data Continuity Reason": data_continuity_reason,
        "Gap Quality Score": round(gap_quality_score, 1),
        "Gap Quality Reason": gap_quality_reason,
        "Runtime Confidence Score": runtime_confidence_score,
        "Runtime Confidence Reason": runtime_reason,
        "Replacement Quality Score": round(replacement_quality_score, 1),
        "Replacement Quality Reason": replacement_reason,
        "Prediction Stability Score": prediction_stability_score,
        "Prediction Stability Reason": prediction_stability_reason,
        "Data Quality Score": data_quality_score,
        "Data Quality Reason": data_quality_reason,
        "Overall Explanation": explanation,
        "Algorithm Version": ALGORITHM_VERSION,
    }

    trend = {
        "Run Date": run_date,
        "Hospital Name": "",
        "Serial Number": "",
        "Component Type": "Degas Module",
        "Forecast Readiness": readiness,
        "Overall Confidence Score": overall,
        "Overall Confidence Stars": stars,
        "Coverage Score": coverage_score,
        "Data Continuity Score": round(data_continuity_score, 1),
        "Gap Quality Score": round(gap_quality_score, 1),
        "Runtime Confidence Score": runtime_confidence_score,
        "Replacement Quality Score": round(replacement_quality_score, 1),
        "Prediction Stability Score": prediction_stability_score,
        "Data Quality Score": data_quality_score,
        "Measured Runtime Hours": measured,
        "Estimated Runtime Hours": estimated,
        "Unknown Days": unknown_days,
        "Algorithm Version": ALGORITHM_VERSION,
    }

    return [row], [trend]


def save_explainable_reliability(summary_rows, runtime_rows, gap_rows, replacement_validation_rows=None, prediction_rows=None):
    rows, trend_rows = build_explainable_reliability(summary_rows, runtime_rows, gap_rows, replacement_validation_rows, prediction_rows)
    p1 = append_explainable_rows("Explainable_Reliability.xlsx", EXPLAINABLE_RELIABILITY_HEADERS, rows)
    p2 = append_explainable_rows("Reliability_Trend.xlsx", RELIABILITY_TREND_HEADERS, trend_rows)
    return (p1, p2), rows, trend_rows

def build_prediction_and_confidence(runtime_rows):
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    observed = round(sum(float(r.get("Measured Runtime Hours", 0) or 0) for r in runtime_rows), 2)
    estimated = round(sum(float(r.get("Estimated Runtime Hours", 0) or 0) for r in runtime_rows), 2)
    unknown_days = sum(int(r.get("Unknown Days", 0) or 0) for r in runtime_rows)
    total = round(observed + estimated, 2)

    coverage_score = 100 if observed > 0 else 0
    gap_score = max(0, 100 - unknown_days * 3)
    replacement_score = 100
    runtime_quality_score = 100 if estimated == 0 else max(50, 100 - min(50, (estimated / max(1, total)) * 100))
    prediction_stability_score = 50
    overall = round(coverage_score*0.20 + gap_score*0.20 + replacement_score*0.20 + runtime_quality_score*0.25 + prediction_stability_score*0.15, 1)

    pred = [{
        "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
        "Component Type": "Degas Module", "Observed Runtime Hours": observed,
        "Estimated Runtime Hours": estimated, "Total Runtime Basis Hours": total,
        "Forecast Method": "Runtime reconstruction foundation",
        "Prediction Status": "Not calculated",
        "Prediction Note": "Version 7.2c reconstructs runtime basis and saves explainable reliability/readiness. Final replacement prediction date is not calculated yet.",
        "Algorithm Version": ALGORITHM_VERSION,
    }]
    conf = [{
        "Run Date": run_date, "Hospital Name": "", "Serial Number": "",
        "Component Type": "Degas Module", "Coverage Score": coverage_score,
        "Gap Quality Score": gap_score, "Replacement History Score": replacement_score,
        "Runtime Quality Score": round(runtime_quality_score, 1),
        "Prediction Stability Score": prediction_stability_score,
        "Overall Confidence Score": overall, "Overall Confidence Stars": confidence_stars(overall),
        "Explanation": "Observed %.2f h + estimated %.2f h = runtime basis %.2f h. Unknown days: %d." % (observed, estimated, total, unknown_days),
        "Algorithm Version": ALGORITHM_VERSION,
    }]
    return pred, conf

def save_runtime_reconstruction(summary_rows, missing_handling="Unknown"):
    runtime_rows, gap_rows = build_runtime_reconstruction(summary_rows, missing_handling)
    prediction_rows, confidence_rows = build_prediction_and_confidence(runtime_rows)
    replacement_validation_rows = validate_replacement_history() if "validate_replacement_history" in globals() else []
    explain_paths, explain_rows, trend_rows = save_explainable_reliability(summary_rows, runtime_rows, gap_rows, replacement_validation_rows, prediction_rows)
    p1 = append_runtime_rows("Runtime_Reconstruction.xlsx", RUNTIME_RECONSTRUCTION_HEADERS, runtime_rows)
    p2 = append_runtime_rows("Gap_History.xlsx", GAP_HISTORY_HEADERS, gap_rows)
    p3 = append_runtime_rows("Prediction_History.xlsx", PREDICTION_HISTORY_HEADERS, prediction_rows)
    p4 = append_runtime_rows("Confidence_History.xlsx", CONFIDENCE_HISTORY_HEADERS, confidence_rows)
    return (p1, p2, p3, p4) + explain_paths, runtime_rows, gap_rows, prediction_rows, confidence_rows

class RuntimeReconstructionDialog(QtWidgets.QDialog):
    def __init__(self, runtime_rows, gap_rows, prediction_rows, confidence_rows, parent=None):
        super(RuntimeReconstructionDialog, self).__init__(parent)
        self.setWindowTitle("Runtime Reconstruction")
        self.resize(1250, 720)
        self.runtime_rows = runtime_rows or []
        self.gap_rows = gap_rows or []
        self.prediction_rows = prediction_rows or []
        self.confidence_rows = confidence_rows or []
        self._ui()

    def _ui(self):
        layout = QVBoxLayout(self)
        if self.prediction_rows:
            p = self.prediction_rows[-1]
            c = self.confidence_rows[-1] if self.confidence_rows else {}
            summary = (
                "<b>Runtime Reconstruction Summary</b><br>"
                "Observed Runtime: <b>{obs} h</b> / Estimated Runtime: <b>{est} h</b> / "
                "Runtime Basis: <b>{total} h</b><br>"
                "Overall Confidence: <b>{stars}</b> ({score})<br>"
                "{note}"
            ).format(
                obs=p.get("Observed Runtime Hours", ""),
                est=p.get("Estimated Runtime Hours", ""),
                total=p.get("Total Runtime Basis Hours", ""),
                stars=c.get("Overall Confidence Stars", ""),
                score=c.get("Overall Confidence Score", ""),
                note=p.get("Prediction Note", ""),
            )
        else:
            summary = "<b>Runtime Reconstruction Summary</b><br>No data."
        label = QLabel(summary)
        label.setWordWrap(True)
        layout.addWidget(label)

        tabs = QtWidgets.QTabWidget()
        layout.addWidget(tabs, 1)
        self.add_table_tab(tabs, "Runtime Reconstruction", RUNTIME_RECONSTRUCTION_HEADERS, self.runtime_rows)
        self.add_table_tab(tabs, "Gap History", GAP_HISTORY_HEADERS, self.gap_rows)
        self.add_table_tab(tabs, "Prediction History", PREDICTION_HISTORY_HEADERS, self.prediction_rows)
        self.add_table_tab(tabs, "Confidence History", CONFIDENCE_HISTORY_HEADERS, self.confidence_rows)

        b = QPushButton("Close")
        b.clicked.connect(self.accept)
        layout.addWidget(b)

    def add_table_tab(self, tabs, title, headers, rows):
        table = QtWidgets.QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, h in enumerate(headers):
                table.setItem(r_idx, c_idx, QtWidgets.QTableWidgetItem(str(row.get(h, ""))))
        table.resizeColumnsToContents()
        tabs.addTab(table, title)

class ReliabilityDashboardDialog(QtWidgets.QDialog):
    def __init__(self, validation_rows, gap_rows, reliability_rows, snapshot_rows, parent=None):
        super(ReliabilityDashboardDialog, self).__init__(parent)
        self.setWindowTitle("Reliability Dashboard")
        self.resize(1200, 720)
        self.validation_rows = validation_rows or []
        self.gap_rows = gap_rows or []
        self.reliability_rows = reliability_rows or []
        self.snapshot_rows = snapshot_rows or []
        self._ui()

    def _ui(self):
        layout = QVBoxLayout(self)
        if self.reliability_rows:
            r = self.reliability_rows[-1]
            summary = (
                "<b>Reliability Summary</b><br>"
                "Coverage: <b>{coverage}%</b> / Confidence: <b>{stars}</b> ({score}) / "
                "Data Quality: <b>{quality}</b><br>"
                "Estimated Runtime: <b>{est} h</b> / Warning: {warning}"
            ).format(
                coverage=r.get("Coverage %", ""),
                stars=r.get("Confidence Stars", ""),
                score=r.get("Confidence Score", ""),
                quality=r.get("Data Quality", ""),
                est=r.get("Estimated Runtime Hours", ""),
                warning=r.get("Warning", ""),
            )
        else:
            summary = "<b>Reliability Summary</b><br>No reliability data."
        label = QLabel(summary)
        label.setWordWrap(True)
        layout.addWidget(label)

        tabs = QtWidgets.QTabWidget()
        layout.addWidget(tabs, 1)

        self.add_table_tab(tabs, "Reliability", RELIABILITY_HISTORY_HEADERS, self.reliability_rows)
        self.add_table_tab(tabs, "Replacement Validation", REPLACEMENT_VALIDATION_HEADERS, self.validation_rows)
        self.add_table_tab(tabs, "Gap Analysis", GAP_ANALYSIS_HEADERS, self.gap_rows)
        self.add_table_tab(tabs, "Forecast Snapshot", FORECAST_SNAPSHOT_HEADERS, self.snapshot_rows)

        b = QPushButton("Close")
        b.clicked.connect(self.accept)
        layout.addWidget(b)

    def add_table_tab(self, tabs, title, headers, rows):
        table = QtWidgets.QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, h in enumerate(headers):
                table.setItem(r_idx, c_idx, QtWidgets.QTableWidgetItem(str(row.get(h, ""))))
        table.resizeColumnsToContents()
        tabs.addTab(table, title)

DATA_COVERAGE_HEADERS = ["Run Date","Hospital Name","Serial Number","Component Type","First Log Date","Last Log Date","Observation Days","Missing Log Handling","Coverage %","Data Quality","Confidence","Warning","Algorithm Version"]
FORECAST_RUN_HISTORY_HEADERS = ["Run Date","Hospital Name","Serial Number","Component Type","Missing Log Handling","Replacement Count","First Log Date","Last Log Date","Coverage %","Data Quality","Confidence","Forecast Basis","Warning","Algorithm Version"]

def ensure_forecast_foundation_files():
    for fname, headers in {"Data_Coverage_Master.xlsx": DATA_COVERAGE_HEADERS, "Forecast_Run_History.xlsx": FORECAST_RUN_HISTORY_HEADERS}.items():
        path = master_path(fname)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(fname)[0]
            for c, h in enumerate(headers, 1):
                ws.cell(1, c, h)
                ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 3)
            ws.freeze_panes = "A2"
            wb.save(path)

def append_forecast_rows(path, headers, rows):
    ensure_forecast_foundation_files()
    wb = load_workbook(path)
    ws = wb.active
    for row in rows:
        r = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    wb.save(path)

def save_forecast_foundation(summary_rows, missing_handling="Unknown"):
    ensure_forecast_foundation_files()
    dates = [r.get("DateTime") for r in (summary_rows or []) if isinstance(r.get("DateTime"), dt.datetime)]
    first = min(dates) if dates else None
    last = max(dates) if dates else None
    days = (last - first).days + 1 if first and last else 0
    quality = "High" if days >= 180 else ("Medium" if days >= 30 else "Low")
    confidence = "Medium" if days > 0 else "Low"
    warning = "" if days > 0 else "No log date data available."
    run_date = dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    row = {"Run Date": run_date, "Hospital Name": "", "Serial Number": "", "Component Type": "Degas Module", "First Log Date": first.strftime("%Y/%m/%d") if first else "", "Last Log Date": last.strftime("%Y/%m/%d") if last else "", "Observation Days": days, "Missing Log Handling": missing_handling, "Coverage %": 100.0 if days > 0 else 0.0, "Data Quality": quality, "Confidence": confidence, "Warning": warning, "Algorithm Version": ALGORITHM_VERSION}
    hist = {"Run Date": run_date, "Hospital Name": "", "Serial Number": "", "Component Type": "Degas Module", "Missing Log Handling": missing_handling, "Replacement Count": "", "First Log Date": row["First Log Date"], "Last Log Date": row["Last Log Date"], "Coverage %": row["Coverage %"], "Data Quality": quality, "Confidence": confidence, "Forecast Basis": "Coverage foundation only / no final prediction in Version 7.0 Initial", "Warning": warning, "Algorithm Version": ALGORITHM_VERSION}
    append_forecast_rows(master_path("Data_Coverage_Master.xlsx"), DATA_COVERAGE_HEADERS, [row])
    append_forecast_rows(master_path("Forecast_Run_History.xlsx"), FORECAST_RUN_HISTORY_HEADERS, [hist])
    return [row], [hist]

class ForecastFoundationDialog(QtWidgets.QDialog):
    def __init__(self, coverage_rows, history_rows, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Forecast Foundation / Data Coverage")
        self.resize(1000, 600)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Version 7.0 Initial: Forecast foundation only. Final prediction is not calculated yet."))
        tabs = QtWidgets.QTabWidget()
        layout.addWidget(tabs, 1)
        for title, headers, rows in [("Data Coverage", DATA_COVERAGE_HEADERS, coverage_rows), ("Forecast Run History", FORECAST_RUN_HISTORY_HEADERS, history_rows)]:
            table = QtWidgets.QTableWidget()
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(len(rows))
            for r_idx, row in enumerate(rows):
                for c_idx, h in enumerate(headers):
                    table.setItem(r_idx, c_idx, QtWidgets.QTableWidgetItem(str(row.get(h, ""))))
            table.resizeColumnsToContents()
            tabs.addTab(table, title)
        b = QPushButton("Close")
        b.clicked.connect(self.accept)
        layout.addWidget(b)

def import_history_xlsx_to_master(path):
    """
    Import the attached History.xlsx layout:
      Row 1: SN / HP
      Column A: HP SN
      Column B: HP/Site name
      Columns C onward: replacement dates as Excel serials or date text

    This step only imports/editable history data.
    It does not calculate forecast, MTBF, reliability, or remaining life.
    """
    ensure_master_foundation_files()

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    hp_existing = read_master_foundation_rows("Hospital_Master.xlsx", HP_MASTER_HEADERS)
    rep_existing = read_master_foundation_rows("Replacement_History.xlsx", REPLACEMENT_HISTORY_HEADERS)

    hp_by_id = {}
    for row in hp_existing:
        key = normalize_history_key(row.get("Hospital Name"))
        if key:
            hp_by_id[key] = row

    rep_keys = set()
    for row in rep_existing:
        hp_id = normalize_history_key(row.get("Hospital Name"))
        d = excel_serial_to_datetime(row.get("Replacement Date"))
        d_key = d.strftime("%Y-%m-%d") if d else str(row.get("Replacement Date", "")).strip()
        rep_keys.add((hp_id, d_key))

    imported_hp = 0
    imported_rep = 0
    skipped_dates = 0

    for r in range(2, ws.max_row + 1):
        hp_id = normalize_history_key(ws.cell(r, 1).value)
        site = normalize_history_key(ws.cell(r, 2).value)

        if not hp_id and not site:
            continue
        if hp_id.upper() == "NA":
            continue

        if hp_id:
            if hp_id not in hp_by_id:
                hp_row = {
                    "Hospital Name": site or hp_id,
                    "Site": site,
                    "Serial Number": hp_id,
                    "Installation Date": "",
                    "Coil Type": "",
                    "Status": "Active",
                    "Comment": "Imported from History.xlsx",
                }
                hp_by_id[hp_id] = hp_row
                hp_existing.append(hp_row)
                imported_hp += 1
            else:
                # Fill missing site/serial without overwriting user edits.
                if site and not hp_by_id[hp_id].get("Site"):
                    hp_by_id[hp_id]["Site"] = site
                if not hp_by_id[hp_id].get("Serial Number"):
                    hp_by_id[hp_id]["Serial Number"] = hp_id

        for c in range(3, ws.max_column + 1):
            raw = ws.cell(r, c).value
            if raw in ("", None):
                continue
            d = excel_serial_to_datetime(raw)
            if not d:
                skipped_dates += 1
                continue

            d_key = d.strftime("%Y-%m-%d")
            key = (hp_id, d_key)
            if key in rep_keys:
                continue

            rep_existing.append({
                "Hospital Name": site or hp_id,
                "Serial Number": hp_id,
                "Component Type": "Degas Module",
                "Component ID": "",
                "Coil ID": "",
                "Replacement Date": d.strftime("%Y/%m/%d"),
                "Replacement Type": "Unknown",
                "Reason": "",
                "Comment": "Imported from History.xlsx",
            })
            rep_keys.add(key)
            imported_rep += 1

    hp_path = write_master_foundation_rows("Hospital_Master.xlsx", HP_MASTER_HEADERS, hp_existing)
    rep_path = write_master_foundation_rows("Replacement_History.xlsx", REPLACEMENT_HISTORY_HEADERS, rep_existing)

    return {
        "hp_path": hp_path,
        "rep_path": rep_path,
        "imported_hp": imported_hp,
        "imported_replacement": imported_rep,
        "skipped_dates": skipped_dates,
        "total_hp": len(hp_existing),
        "total_replacement": len(rep_existing),
    }



class MasterFoundationDialog(QtWidgets.QDialog):
    """
    Step5 simple master editor.
    This is intentionally basic and safe: no Forecast calculation is performed.
    """
    def __init__(self, parent=None):
        super(MasterFoundationDialog, self).__init__(parent)
        self.setWindowTitle("Step6.2 Master Foundation")
        self.resize(1000, 650)
        self.hp_rows = []
        self.rep_rows = []
        self._ui()
        self.load_all()

    def _ui(self):
        layout = QVBoxLayout(self)
        info = QLabel("Step6.2 Master Foundation: edit HP Master and Replacement History, or import History.xlsx. No forecast calculation is performed in this step.")
        info.setWordWrap(True)
        layout.addWidget(info)

        self.tabs = QtWidgets.QTabWidget()
        layout.addWidget(self.tabs, 1)

        self.hp_table = QtWidgets.QTableWidget()
        self.hp_table.setColumnCount(len(HP_MASTER_HEADERS))
        self.hp_table.setHorizontalHeaderLabels(HP_MASTER_HEADERS)
        self.hp_table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.hp_table, "HP Master")

        self.component_table = QtWidgets.QTableWidget()
        self.component_table.setColumnCount(len(COMPONENT_MASTER_HEADERS))
        self.component_table.setHorizontalHeaderLabels(COMPONENT_MASTER_HEADERS)
        self.component_table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.component_table, "Component Master")

        self.rep_table = QtWidgets.QTableWidget()
        self.rep_table.setColumnCount(len(REPLACEMENT_HISTORY_HEADERS))
        self.rep_table.setHorizontalHeaderLabels(REPLACEMENT_HISTORY_HEADERS)
        self.rep_table.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.rep_table, "Replacement History")

        btns = QHBoxLayout()
        layout.addLayout(btns)

        b = QPushButton("Add Row")
        b.clicked.connect(self.add_row)
        btns.addWidget(b)

        b = QPushButton("Delete Selected Row")
        b.clicked.connect(self.delete_selected_row)
        btns.addWidget(b)

        b = QPushButton("Save")
        b.clicked.connect(self.save_all)
        btns.addWidget(b)

        b = QPushButton("Reload")
        b.clicked.connect(self.load_all)
        btns.addWidget(b)

        b = QPushButton("Import History.xlsx")
        b.clicked.connect(self.import_history_file)
        btns.addWidget(b)

        b = QPushButton("Open Master Folder")
        b.clicked.connect(self.open_master_folder)
        btns.addWidget(b)

        b = QPushButton("Close")
        b.clicked.connect(self.accept)
        btns.addWidget(b)

    def show_wait(self, text="Updating... Please wait."):
        try:
            self._wait = QProgressDialog(text, None, 0, 0, self)
            self._wait.setWindowTitle("Updating")
            self._wait.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            self._wait.setMinimumDuration(0)
            self._wait.setCancelButton(None)
            self._wait.show()
            QApplication.processEvents()
        except Exception:
            self._wait = None

    def hide_wait(self):
        try:
            if getattr(self, "_wait", None):
                self._wait.close()
                self._wait = None
            QApplication.processEvents()
        except Exception:
            pass

    def active_table_and_headers(self):
        if self.tabs.currentIndex() == 0:
            return self.hp_table, HP_MASTER_HEADERS, "Hospital_Master.xlsx"
        if self.tabs.currentIndex() == 1:
            return self.component_table, COMPONENT_MASTER_HEADERS, "Component_Master.xlsx"
        return self.rep_table, REPLACEMENT_HISTORY_HEADERS, "Replacement_History.xlsx"

    def populate_table(self, table, headers, rows):
        table.setRowCount(len(rows))
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        for r_idx, row in enumerate(rows):
            for c_idx, h in enumerate(headers):
                val = row.get(h, "")
                if val is None:
                    val = ""
                item = QtWidgets.QTableWidgetItem(str(val))
                table.setItem(r_idx, c_idx, item)
        table.resizeColumnsToContents()

    def table_to_rows(self, table, headers):
        rows = []
        for r in range(table.rowCount()):
            row = {}
            empty = True
            for c, h in enumerate(headers):
                item = table.item(r, c)
                val = item.text().strip() if item else ""
                if val:
                    empty = False
                row[h] = val
            if not empty:
                rows.append(row)
        return rows

    def load_all(self):
        try:
            ensure_master_foundation_files()
            self.hp_rows = read_master_foundation_rows("Hospital_Master.xlsx", HP_MASTER_HEADERS)
            self.component_rows = read_master_foundation_rows("Component_Master.xlsx", COMPONENT_MASTER_HEADERS)
            self.rep_rows = read_master_foundation_rows("Replacement_History.xlsx", REPLACEMENT_HISTORY_HEADERS)
            self.populate_table(self.hp_table, HP_MASTER_HEADERS, self.hp_rows)
            self.populate_table(self.component_table, COMPONENT_MASTER_HEADERS, self.component_rows)
            self.populate_table(self.rep_table, REPLACEMENT_HISTORY_HEADERS, self.rep_rows)
        except Exception:
            self.hide_wait()
            QMessageBox.critical(self, "Master Foundation", traceback.format_exc())

    def save_all(self):
        try:
            self.show_wait("Saving master files...")
            hp_rows = self.table_to_rows(self.hp_table, HP_MASTER_HEADERS)
            component_rows = self.table_to_rows(self.component_table, COMPONENT_MASTER_HEADERS)
            rep_rows = self.table_to_rows(self.rep_table, REPLACEMENT_HISTORY_HEADERS)
            hp_path = write_master_foundation_rows("Hospital_Master.xlsx", HP_MASTER_HEADERS, hp_rows)
            component_path = write_master_foundation_rows("Component_Master.xlsx", COMPONENT_MASTER_HEADERS, component_rows)
            rep_path = write_master_foundation_rows("Replacement_History.xlsx", REPLACEMENT_HISTORY_HEADERS, rep_rows)
            self.hide_wait()
            QMessageBox.information(self, "Master Foundation", "Saved:\n%s\n%s\n%s" % (hp_path, component_path, rep_path))
        except Exception:
            QMessageBox.critical(self, "Master Foundation", traceback.format_exc())

    def add_row(self):
        table, headers, _fname = self.active_table_and_headers()
        r = table.rowCount()
        table.insertRow(r)
        for c in range(len(headers)):
            table.setItem(r, c, QtWidgets.QTableWidgetItem(""))

    def delete_selected_row(self):
        table, _headers, _fname = self.active_table_and_headers()
        rows = sorted(set(i.row() for i in table.selectedIndexes()), reverse=True)
        for r in rows:
            table.removeRow(r)

    def import_history_file(self):
        try:
            path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select History.xlsx", "", "Excel Files (*.xlsx *.xlsm);;All Files (*.*)")
            if not path:
                return
            self.show_wait("Importing History.xlsx... Please wait.")
            result = import_history_xlsx_to_master(path)
            self.load_all()
            self.hide_wait()
            QMessageBox.information(
                self,
                "Master Foundation",
                "History import complete.\n\nImported data was written to Hospital_Master.xlsx and Replacement_History.xlsx.\n\n"
                "Imported HP rows: %d\n"
                "Imported replacement rows: %d\n"
                "Skipped invalid date cells: %d\n\n"
                "HP total: %d\n"
                "Replacement total: %d\n\n"
                "Saved:\n%s\n%s\n\n"
                "Replacement Type is set to Unknown. Edit each row to Failure or Preventive before Forecast calculation."
                % (
                    result["imported_hp"],
                    result["imported_replacement"],
                    result["skipped_dates"],
                    result["total_hp"],
                    result["total_replacement"],
                    result["hp_path"],
                    result["rep_path"],
                )
            )
        except Exception:
            QMessageBox.critical(self, "Master Foundation", traceback.format_exc())

    def open_master_folder(self):
        try:
            ensure_master_foundation_files()
            os.startfile(os.path.dirname(master_foundation_path("Hospital_Master.xlsx")))
        except Exception:
            QMessageBox.critical(self, "Master Foundation", traceback.format_exc())


def ensure_data_folders():
    root = get_data_root()
    for sub in ["Master", "History", "Backup", "Log", "Reports"]:
        p = os.path.join(root, sub)
        if not os.path.exists(p):
            os.makedirs(p)
    ensure_master_files(root)
    return root


def master_path(name):
    return os.path.join(ensure_data_folders(), "Master", name)


def ensure_master_files(root):
    master = os.path.join(root, "Master")
    files = {
        "SN_Master.xlsx": ["SN", "Hospital", "System", "Install Date", "Last Update", "Comment"],
        "Hospital_Master.xlsx": ["Hospital", "Country", "Region", "Installation Date", "Data Coverage Note", "Comment"],
        "Replacement_History.xlsx": ["SN", "Hospital", "Replacement Date", "Replacement Type", "Part", "Reason", "Engineer", "Comment"],
        "Forecast_Master.xlsx": ["Calc Date", "SN", "Hospital", "Forecast Mode", "Forecast Scope", "Installation Date", "Last Replacement Date", "Data Start Date", "Data End Date", "Data Coverage Status", "Used Data Count", "Excluded Data Count", "Forecast Date", "Remaining Days", "Confidence", "Algorithm Version"],
        "Outlier_Master.xlsx": ["SN", "Date", "File", "Item", "Value", "Reason", "User", "Date Removed", "Use for Forecast"],
        "Ignore_Period.xlsx": ["SN", "Start Date", "End Date", "Reason", "User", "Created Date"],
    }
    for name, headers in files.items():
        p = os.path.join(master, name)
        if not os.path.exists(p):
            wb = Workbook()
            ws = wb.active
            ws.title = os.path.splitext(name)[0]
            for c, h in enumerate(headers, 1):
                ws.cell(1, c, h).font = Font(bold=True)
            wb.save(p)
        else:
            ensure_headers(p, headers)


def ensure_headers(path, required_headers):
    try:
        wb = load_workbook(path)
        ws = wb.active
        existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        changed = False
        for h in required_headers:
            if h not in existing:
                ws.cell(1, ws.max_column + 1, h).font = Font(bold=True)
                existing.append(h)
                changed = True
        if changed:
            wb.save(path)
    except Exception:
        pass


def read_master_rows(path):
    if not path or not os.path.exists(path):
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_value = False
        for c, h in enumerate(headers, 1):
            if h:
                v = ws.cell(r, c).value
                row[h] = v
                if v not in ("", None):
                    has_value = True
        if has_value:
            rows.append(row)
    return rows


def get_hospital_for_sn(sn):
    for r in read_master_rows(master_path("SN_Master.xlsx")):
        if str(r.get("SN", "")).strip() == str(sn):
            return str(r.get("Hospital", "") or "").strip()
    return ""


def get_installation_date_for_site(sn, hospital=""):
    """
    Priority:
      1) SN_Master.Install Date for the specific SN
      2) Hospital_Master.Installation Date for the hospital/site
      3) None
    """
    for r in read_master_rows(master_path("SN_Master.xlsx")):
        if str(r.get("SN", "")).strip() == str(sn):
            d = coerce_datetime(r.get("Install Date"))
            if d:
                return d

    hospital = hospital or get_hospital_for_sn(sn)
    if hospital:
        for r in read_master_rows(master_path("Hospital_Master.xlsx")):
            if str(r.get("Hospital", "")).strip() == str(hospital):
                d = coerce_datetime(r.get("Installation Date"))
                if d:
                    return d
    return None


def data_coverage_status(installation_date, last_replacement_date, data_start_date):
    if not data_start_date:
        return "No data"

    base_date = last_replacement_date or installation_date

    if not base_date:
        return "Unknown installation/replacement"

    gap_days = (data_start_date - base_date).days

    if gap_days <= 7:
        return "Full from install/replacement"

    return "Left-censored: no older data before first log (%d days missing)" % gap_days


def forecast_data_start_date(sn, hospital, mode, data_start_date):
    """
    Forecast should start from the later of:
      - last replacement date for the selected mode
      - installation date
      - first available data date

    This handles sites where there is no data before the first log.
    """
    install_date = get_installation_date_for_site(sn, hospital)
    last_rep = None
    if "get_last_replacement_date" in globals():
        try:
            last_rep = get_last_replacement_date(sn, mode)
        except Exception:
            last_rep = None

    candidates = [d for d in [install_date, last_rep, data_start_date] if d]
    return max(candidates) if candidates else data_start_date



def is_clean_state(s):
    return str(s).upper() in ("CLEAN_TANK_CIRCULATE", "CLEAN_XD_CIRCULATE")


def parse_watersystem_log(path):
    treat = degas = tank = xd = dpause = tpause = errtime = 0.0
    clean_count = 0
    clean_active = False
    clean_context = ""

    start = last = end = None
    prev_state = ""
    prev_err = ""

    sensors = {
        "Primary Flow": [0.0, 0],
        "Chiller Temp": [0.0, 0],
        "Absolute Pressure": [0.0, 0],
        "Dynamic Pressure": [0.0, 0],
    }

    raw = open(path, "rb").read()
    lines = []
    for enc in ("utf-8-sig", "cp932", "latin-1"):
        try:
            lines = raw.decode(enc, errors="replace").splitlines()
            break
        except Exception:
            pass

    for line in lines:
        cur = parse_time_seconds(line)
        if cur is None:
            continue

        state = get_token(line, 1).upper()
        er = get_token(line, 3).upper()

        if start is None:
            start = cur
        elif last is not None:
            diff = cur - last
            if diff < 0:
                diff = diff + 86400 if last > 86300 and cur < 300 else 0

            if prev_state == "TREAT_CIRCULATE":
                treat += diff
            elif prev_state == "DEGAS_CIRCULATE":
                degas += diff
            elif prev_state == "CLEAN_TANK_CIRCULATE":
                tank += diff
            elif prev_state == "CLEAN_XD_CIRCULATE":
                xd += diff
            elif prev_state == "DEGAS_PAUSE":
                dpause += diff
            elif prev_state == "TREAT_PAUSE":
                tpause += diff

            if prev_err and prev_err != "NO_ERROR":
                errtime += diff

            if clean_context and not is_clean_state(prev_state):
                if clean_context == "CLEAN_TANK_CIRCULATE":
                    tank += diff
                elif clean_context == "CLEAN_XD_CIRCULATE":
                    xd += diff

        if "DRAIN" in state:
            clean_active = False
            clean_context = ""
        elif is_clean_state(state):
            if not clean_active:
                clean_count += 1
                clean_active = True
            clean_context = state

        for label, keys in {
            "Primary Flow": ["PrimaryFlow", "Primary Flow", "PrimaryFlowMeter"],
            "Chiller Temp": ["ChillerTemp", "Chiller Temp"],
            "Absolute Pressure": ["AbsolutePressure", "Absolute Pressure"],
            "Dynamic Pressure": ["DynamicPressure", "Dynamic Pressure"],
        }.items():
            for k in keys:
                m = re.search(re.escape(k) + r"\s*[:=]?\s*(-?\d+(?:\.\d+)?)", line, re.I)
                if m:
                    sensors[label][0] += float(m.group(1))
                    sensors[label][1] += 1
                    break

        last = end = cur
        prev_state = state
        prev_err = er

    if start is None:
        return None

    runtime = (end or 0) - start
    if runtime < 0:
        runtime += 86400

    if clean_count == 0 and (tank + xd) > 0:
        clean_count = 1

    h = file_crc32_hex(path)
    size = os.path.getsize(path)

    def avg(label):
        total, cnt = sensors[label]
        return total / cnt if cnt else ""

    return {
        "File Name": os.path.basename(path),
        "Start Datetime (YYYY-MM-DD HH:MM)": build_file_datetime_text(path),
        "TREAT_CIRCULATE (HH:MM:SS)": seconds_to_hms(treat),
        "DEGAS_CIRCULATE (HH:MM:SS)": seconds_to_hms(degas),
        "CLEAN_TANK_CIRCULATE (HH:MM:SS)": seconds_to_hms(tank),
        "CLEAN_XD_CIRCULATE (HH:MM:SS)": seconds_to_hms(xd),
        "Total File Runtime (HH:MM:SS)": seconds_to_hms(runtime),
        "Total File Runtime (Seconds)": runtime,
        "Total Circulate Time (Seconds)": treat + degas + tank + xd,
        "DEGAS_PAUSE (HH:MM:SS)": seconds_to_hms(dpause),
        "TREAT_PAUSE (HH:MM:SS)": seconds_to_hms(tpause),
        "ERROR (HH:MM:SS)": seconds_to_hms(errtime),
        "Clean Count": clean_count,
        "Primary Flow": avg("Primary Flow"),
        "Chiller Temp": avg("Chiller Temp"),
        "Absolute Pressure": avg("Absolute Pressure"),
        "Dynamic Pressure": avg("Dynamic Pressure"),
        "File CRC32": h,
        "File Size": size,
    }


def find_latest_result_summary(folder):
    files = glob.glob(os.path.join(folder, "ResultSummary*.xlsx"))
    return sorted(files, key=os.path.getmtime, reverse=True)[0] if files else None


def read_existing_rows(path, progress=None, cancel=None):
    rows = []
    keys = set()

    if not path or not os.path.exists(path):
        return rows, keys

    wb = load_workbook(path, data_only=True, read_only=True)
    if "Result" not in wb.sheetnames:
        return rows, keys

    ws = wb["Result"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [h or "" for h in header_row]
    max_row = ws.max_row or 1

    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if cancel and cancel():
            break
        if not values or not values[0]:
            continue

        row = {}
        for i, h in enumerate(headers):
            if h and i < len(values):
                row[h] = values[i]

        row["Start Datetime (YYYY-MM-DD HH:MM)"] = coerce_datetime(row.get("Start Datetime (YYYY-MM-DD HH:MM)"))
        rows.append(row)

        keys.add(result_key(row.get("File Name", ""), row.get("Start Datetime (YYYY-MM-DD HH:MM)")))
        if row.get("File Name"):
            keys.add(name_key(row.get("File Name")))
        ck = content_key(row.get("File CRC32"), row.get("File Size"))
        if ck:
            keys.add(ck)

        if progress and (idx == 2 or idx % 250 == 0 or idx == max_row):
            try:
                progress(idx, max_row, "Reading existing ResultSummary rows")
            except Exception:
                pass

    try:
        wb.close()
    except Exception:
        pass
    return rows, keys



def read_chart_summary_direct(path):
    """
    Load chart data directly from Daily_Summary when available.
    This fixes cases where ResultSummary exists but the Result sheet cannot be
    rebuilt into chart rows because date/time column names differ.
    """
    if not path or not os.path.exists(path):
        return []

    wb = load_workbook(path, data_only=True, read_only=True)

    target_sheet = None
    for name in ["Daily_Summary", "Daily Summary", "Summary", "Chart_Data", "Chart Data"]:
        if name in wb.sheetnames:
            target_sheet = wb[name]
            break

    if target_sheet is None:
        return []

    ws = target_sheet
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    header_map = {h.lower(): i + 1 for i, h in enumerate(headers) if h}

    def get_cell(row, names):
        for n in names:
            c = header_map.get(n.lower())
            if c:
                return ws.cell(row, c).value
        return None

    def num(v):
        try:
            if v in ("", None):
                return 0.0
            return float(v)
        except Exception:
            return 0.0

    out = []
    for r in range(2, ws.max_row + 1):
        raw_dt = get_cell(r, ["DateTime", "Date Time", "Start Datetime (YYYY-MM-DD HH:MM)", "Start Datetime", "Datetime"])
        d = coerce_datetime(raw_dt)
        if not d:
            # Try date + time split columns.
            raw_date = get_cell(r, ["Date", "Start Date"])
            raw_time = get_cell(r, ["Time", "Start Time"])
            dd = coerce_datetime(raw_date)
            if dd and raw_time not in ("", None):
                try:
                    if isinstance(raw_time, dt.time):
                        d = dt.datetime.combine(dd.date(), raw_time)
                    else:
                        ts = str(raw_time).strip()
                        for fmt in ["%H:%M:%S", "%H:%M"]:
                            try:
                                tt = dt.datetime.strptime(ts, fmt).time()
                                d = dt.datetime.combine(dd.date(), tt)
                                break
                            except Exception:
                                pass
                except Exception:
                    d = dd
            else:
                d = dd

        if not d:
            continue

        row = {"DateTime": d}
        for key in [
            "Treat", "Degas", "Clean", "Runtime",
            "DEGAS_PAUSE", "TREAT_PAUSE", "ERROR",
            "Treat Cum", "Degas Cum", "Clean Cum", "Runtime Cum",
            "DEGAS_PAUSE Cum", "TREAT_PAUSE Cum", "ERROR Cum",
        ]:
            row[key] = num(get_cell(r, [key, key.replace(" ", ""), key.replace("_", " ")]))
        out.append(row)

    return out


def collect_files(folder, progress=None, cancel=None):
    files = []
    scanned = 0
    errors = []
    for root, dirs, names in os.walk(folder):
        if cancel and cancel():
            break
        dirs[:] = [d for d in dirs if d.lower() not in ("$recycle.bin", "system volume information", "__pycache__")]
        for name in names:
            if cancel and cancel():
                break
            scanned += 1
            p = os.path.join(root, name)
            try:
                if is_watersystem_file(p):
                    files.append(p)
            except Exception as e:
                if len(errors) < 20:
                    errors.append("%s : %s" % (p, e))
            if progress and scanned % 100 == 0:
                progress(scanned, os.path.basename(root))
    files.sort(key=lambda p: (build_file_datetime_text(os.path.basename(p)) or dt.datetime.min, p.lower()))
    collect_files.last_scan = {"scanned": scanned, "matched": len(files), "errors": errors, "first_files": files[:10]}
    return files


def build_summary_rows(rows):
    """
    Build chart rows from Result rows.

    ChartDataFix:
    - If Start Datetime is missing, fallback to File Name date parsing.
    - If Total File Runtime (Seconds) is missing, fallback to HH:MM:SS.
    - This prevents a blank chart when Result rows exist but Daily_Summary rows were not generated.
    """
    out = []
    cum = dict((k, 0.0) for k in ["Treat", "Degas", "Clean", "Runtime", "DEGAS_PAUSE", "TREAT_PAUSE", "ERROR"])

    def get_dt(r):
        xdt = coerce_datetime(r.get("Start Datetime (YYYY-MM-DD HH:MM)"))
        if xdt:
            return xdt
        fn = r.get("File Name", "")
        if fn:
            return build_file_datetime_text(fn)
        return None

    def runtime_hours(r):
        try:
            sec = float(r.get("Total File Runtime (Seconds)") or 0)
            if sec > 0:
                return sec / 3600.0
        except Exception:
            pass
        return hms_to_seconds(r.get("Total File Runtime (HH:MM:SS)")) / 3600.0

    sorted_rows = []
    for r in rows:
        xdt = get_dt(r)
        if xdt:
            sorted_rows.append((xdt, r))
    sorted_rows.sort(key=lambda x: x[0])

    for xdt, r in sorted_rows:
        vals = {
            "Treat": hms_to_seconds(r.get("TREAT_CIRCULATE (HH:MM:SS)")) / 3600.0,
            "Degas": hms_to_seconds(r.get("DEGAS_CIRCULATE (HH:MM:SS)")) / 3600.0,
            "Clean": (hms_to_seconds(r.get("CLEAN_TANK_CIRCULATE (HH:MM:SS)")) +
                      hms_to_seconds(r.get("CLEAN_XD_CIRCULATE (HH:MM:SS)"))) / 3600.0,
            "Runtime": runtime_hours(r),
            "DEGAS_PAUSE": hms_to_seconds(r.get("DEGAS_PAUSE (HH:MM:SS)")) / 3600.0,
            "TREAT_PAUSE": hms_to_seconds(r.get("TREAT_PAUSE (HH:MM:SS)")) / 3600.0,
            "ERROR": hms_to_seconds(r.get("ERROR (HH:MM:SS)")) / 3600.0,
        }

        for k, v in vals.items():
            cum[k] += v

        row = {"DateTime": xdt}
        row.update(vals)
        row.update({
            "Treat Cum": cum["Treat"],
            "Degas Cum": cum["Degas"],
            "Clean Cum": cum["Clean"],
            "Runtime Cum": cum["Runtime"],
            "DEGAS_PAUSE Cum": cum["DEGAS_PAUSE"],
            "TREAT_PAUSE Cum": cum["TREAT_PAUSE"],
            "ERROR Cum": cum["ERROR"],
        })
        out.append(row)

    return out


def write_resultsummary_workbook(path, rows, summary_rows):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Result")
    for c, h in enumerate(RESULT_HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    for rr, row in enumerate(rows, 2):
        for c, h in enumerate(RESULT_HEADERS, 1):
            ws.cell(rr, c, row.get(h, ""))

    ws.column_dimensions["R"].hidden = True
    ws.column_dimensions["S"].hidden = True
    ws.freeze_panes = "A2"

    for col in range(1, len(RESULT_HEADERS) + 1):
        letter = get_column_letter(col)
        if letter in ("R", "S"):
            continue
        ws.column_dimensions[letter].width = min(max(len(str(RESULT_HEADERS[col - 1])) + 2, 12), 34)

    sw = wb.create_sheet("Daily_Summary")
    if summary_rows:
        headers = list(summary_rows[0].keys())
        for c, h in enumerate(headers, 1):
            sw.cell(1, c, h).font = Font(bold=True)
        for rr, row in enumerate(summary_rows, 2):
            for c, h in enumerate(headers, 1):
                sw.cell(rr, c, row[h])
        sw.freeze_panes = "A2"

    wb.save(path)


def save_or_update_resultsummary(folder, selected_path, rows, summary_rows, serial_number):
    """
    If selected_path exists, update that file.
    If serial_number implies a different file name, save to the new name and remove the old selected file.
    Return the final path.
    """
    final_path = selected_path

    if selected_path:
        current_folder = os.path.dirname(selected_path)
        target_path = result_summary_path_for_sn(current_folder, serial_number)
        if os.path.abspath(target_path).lower() != os.path.abspath(selected_path).lower():
            final_path = target_path
        else:
            final_path = selected_path
    else:
        final_path = result_summary_path_for_sn(folder, serial_number)
        if os.path.exists(final_path):
            final_path = os.path.join(folder, "ResultSummary_%s_%s.xlsx" % (
                safe_sn(serial_number), dt.datetime.now().strftime("%Y%m%d_%H%M%S")))

    # Step2: backup existing ResultSummary before overwrite.
    try:
        if final_path and os.path.exists(final_path):
            backup_dir = os.path.join(os.path.dirname(final_path), "ResultSummary_Backup")
            os.makedirs(backup_dir, exist_ok=True)
            backup_path = os.path.join(
                backup_dir,
                os.path.splitext(os.path.basename(final_path))[0] + "_backup_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
            )
            shutil.copy2(final_path, backup_path)
    except Exception:
        pass

    write_resultsummary_workbook(final_path, rows, summary_rows)

    if selected_path and os.path.exists(selected_path):
        if os.path.abspath(final_path).lower() != os.path.abspath(selected_path).lower():
            try:
                os.remove(selected_path)
            except Exception:
                pass

    return final_path


def calculate_forecast(summary_rows, sn, mode="Failure only", scope="Current SN", selected_hospitals=None):
    hospital = get_hospital_for_sn(sn)
    dates = [coerce_datetime(r.get("DateTime")) for r in summary_rows if coerce_datetime(r.get("DateTime"))]
    raw_start = min(dates) if dates else None
    raw_end = max(dates) if dates else None
    install = get_installation_date_for_site(sn, hospital)
    last_rep = None
    effective_start = forecast_data_start_date(sn, hospital, mode, raw_start)
    coverage = data_coverage_status(install, last_rep, raw_start)
    usable = [r for r in summary_rows if coerce_datetime(r.get("DateTime")) and (not effective_start or coerce_datetime(r.get("DateTime")) >= effective_start)]
    conf = min(90, max(0, len(usable) * 5))
    if "Left-censored" in coverage:
        conf = max(0, conf - 10)
    return {"hospital": hospital, "installation_date": install, "last_replacement": last_rep, "data_start": effective_start, "data_end": raw_end, "coverage_status": coverage, "used": len(usable), "excluded": len(summary_rows)-len(usable), "confidence": conf, "remaining_days": "", "forecast_date": ""}


def append_forecast_master(sn, summary_rows, mode="Failure only", scope="Current SN", selected_hospitals=None):
    return calculate_forecast(summary_rows, sn, mode, scope, selected_hospitals)


class NativeChartView(QGraphicsView):
    """
    Ver40.4 chart logic, rendered with pure Qt.
    No matplotlib / numpy / pandas.
    """

    DAILY_SERIES = ["Treat", "Degas", "Clean", "Runtime", "DEGAS_PAUSE", "TREAT_PAUSE", "ERROR"]
    CUM_SERIES = ["Treat Cum", "Degas Cum", "Clean Cum", "Runtime Cum", "DEGAS_PAUSE Cum", "TREAT_PAUSE Cum", "ERROR Cum"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.scene_obj = QGraphicsScene(self)
        self.setScene(self.scene_obj)
        self.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing, True)
        self.setMinimumSize(760, 520)
        self.summary_rows = []
        self.checks = {}
        self.x0 = None
        self.x1 = None
        self.c0 = None
        self.c1 = None
        self.last_status = ""
        self.manual_x0 = None
        self.manual_x1 = None
        self.current_x0 = None
        self.current_x1 = None
        self.plot_left = 0
        self.plot_top = 0
        self.plot_pw = 1
        self.plot_ph = 1
        self.drag_start = None
        self.drag_item = None
        self.pan_start = None
        self.pan_x0 = None
        self.pan_x1 = None
        self.setMouseTracking(True)
        self.tooltip_item = None
        self.tooltip_bg = None
        self.crosshair_v = None
        self.crosshair_h = None
        self.tooltip_mode = "Point"
        self.last_plot_rows = []
        self.last_active_daily = []
        self.last_active_cum = []
        self.last_cum_values = {}

    def set_data(self, summary_rows, checks, x0, x1, c0, c1):
        self.summary_rows = summary_rows or []
        self.checks = checks or {}
        self.x0 = x0
        self.x1 = x1
        # Keep manual zoom across checkbox/series redraws. Reset is done by double-click.
        self.c0 = c0
        self.c1 = c1
        QtCore.QTimer.singleShot(0, self.redraw)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        QtCore.QTimer.singleShot(0, self.redraw)

    def f(self, value):
        try:
            if value in ("", None):
                return 0.0
            return float(value)
        except Exception:
            return 0.0

    def txt(self, x, y, s, size=9, bold=False, center=False, rotate=0):
        item = self.scene_obj.addText(str(s))
        font = QtGui.QFont()
        font.setPointSize(size)
        font.setBold(bold)
        item.setFont(font)
        item.setDefaultTextColor(QtGui.QColor(40, 40, 40))
        br = item.boundingRect()
        if center:
            item.setPos(x - br.width()/2, y - br.height()/2)
        else:
            item.setPos(x, y - br.height()/2)
        if rotate:
            item.setTransformOriginPoint(br.width()/2, br.height()/2)
            item.setRotation(rotate)

    def message(self, s):
        self.scene_obj.clear()
        w = max(700, self.viewport().width() - 8)
        h = max(480, self.viewport().height() - 8)
        self.scene_obj.setSceneRect(0, 0, w, h)
        self.txt(w/2, h/2 - 18, s, 12, True, True)
        self.txt(w/2, h/2 + 14, "Rows=%d" % len(self.summary_rows), 9, False, True)
        self.last_status = s

    def axis_ranges_for_visible(self, plot_rows, visible_rows, active_daily, active_cum, cum_values):
        """
        Version 7.1a axis rule:
        - Left axis is Daily/bar series. If any daily series is visible, left Y minimum is fixed at 0.
        - Right axis is Cumulative series. It floats independently to the visible cumulative range.
        - If only cumulative is visible, only the right scale matters and it floats.
        """
        daily_vals = []
        for r in plot_rows:
            for n in active_daily:
                daily_vals.append(self.f(r.get(n, 0.0)))
        if daily_vals:
            daily_max = max(daily_vals + [1.0])
            daily_min = 0.0
            daily_max = daily_max * 1.10
        else:
            daily_min, daily_max = 0.0, 1.0

        cum_vals = []
        for n in active_cum:
            cum_vals.extend([self.f(v) for v in cum_values.get(n, {}).values()])
        if cum_vals:
            cum_min = min(cum_vals)
            cum_max = max(cum_vals)
            if cum_min == cum_max:
                margin = max(1.0, abs(cum_max) * 0.05)
            else:
                margin = (cum_max - cum_min) * 0.08
            cum_min -= margin
            cum_max += margin
        else:
            cum_min, cum_max = 0.0, 1.0

        return daily_min, daily_max, cum_min, cum_max

    def redraw(self):
        self.scene_obj.clear()
        w = max(700, self.viewport().width() - 8)
        h = max(480, self.viewport().height() - 8)
        self.scene_obj.setSceneRect(0, 0, w, h)

        if not self.summary_rows:
            self.message("No chart data loaded")
            return

        all_x = [r.get("DateTime") for r in self.summary_rows if isinstance(r.get("DateTime"), dt.datetime)]
        if not all_x:
            self.message("No DateTime data for chart")
            return

        x0 = self.manual_x0 or self.x0 or min(all_x)
        x1 = self.manual_x1 or self.x1 or max(all_x)
        if x1 <= x0:
            x0, x1 = min(all_x), max(all_x)

        visible_rows = [r for r in self.summary_rows if isinstance(r.get("DateTime"), dt.datetime) and x0 <= r["DateTime"] <= x1]
        if not visible_rows:
            self.message("No data in X Start/End range")
            return

        active_daily = [n for n in self.DAILY_SERIES if self.checks.get(n, False)]
        active_cum = [n for n in self.CUM_SERIES if self.checks.get(n, False)]
        self.last_active_daily = active_daily
        self.last_active_cum = active_cum

        if not active_daily and not active_cum:
            self.message("All chart series are OFF")
            return

        # Downsample for UI performance only. Cumulative base still uses all rows.
        if len(visible_rows) > 180:
            step = max(1, len(visible_rows) // 180)
            plot_rows = visible_rows[::step]
            if visible_rows[-1] not in plot_rows:
                plot_rows.append(visible_rows[-1])
        else:
            plot_rows = visible_rows
        self.last_plot_rows = plot_rows

        left, right, top, bottom = 88, 82, 42, 96
        pw = max(100, w - left - right)
        ph = max(100, h - top - bottom)
        self.current_x0 = x0
        self.current_x1 = x1
        self.plot_left = left
        self.plot_top = top
        self.plot_pw = pw
        self.plot_ph = ph

        axis_pen = QtGui.QPen(QtGui.QColor(45,45,45), 1.2)
        grid_pen = QtGui.QPen(QtGui.QColor(225,225,225), 1)
        self.scene_obj.addLine(left, top, left, top+ph, axis_pen)
        self.scene_obj.addLine(left, top+ph, left+pw, top+ph, axis_pen)
        self.scene_obj.addLine(left+pw, top, left+pw, top+ph, axis_pen)
        for i in range(1,5):
            y = top + ph*i/5
            self.scene_obj.addLine(left, y, left+pw, y, grid_pen)

        # Version 7.0: build cumulative values and calculate visible Y axis.
        cum_values = {}
        for n in active_cum:
            vals = []
            running = 0.0
            for r in visible_rows:
                running += self.f(r.get(n.replace(" Cum", ""), 0.0))
                vals.append((r["DateTime"], running))
            cum_values[n] = {d: v for d, v in vals}

        daily_min, daily_max, cum_min, cum_max = self.axis_ranges_for_visible(plot_rows, visible_rows, active_daily, active_cum, cum_values)
        self.last_cum_values = cum_values

        for i in range(6):
            y = top + ph - ph*i/5
            daily_val = daily_min + (daily_max - daily_min) * i / 5.0
            cum_val = cum_min + (cum_max - cum_min) * i / 5.0
            self.txt(left-25, y, "%.1f" % daily_val, 8, False, True)
            self.txt(left+pw+30, y, "%.1f" % cum_val, 8, False, True)
            x = left + pw*i/5
            d = x0 + (x1-x0)*(i/5)
            self.txt(x, top+ph+26, d.strftime("%Y/%m/%d"), 8, False, True, -35)

        self.txt(16, top+ph/2, "Daily Hours", 9, False, True, -90)
        self.txt(left+pw+56, top+ph/2, "Cumulative Hours", 9, False, True, 90)

        span = max(1.0, (x1-x0).total_seconds())
        def xp(x):
            return left + ((x-x0).total_seconds()/span)*pw
        def yd(v):
            return top + ph - ((v - daily_min) / max(0.000001, (daily_max - daily_min))) * ph
        def yc(v):
            return top + ph - ((v - cum_min) / max(0.000001, (cum_max - cum_min))) * ph

        bar_colors = [
            QtGui.QColor(70,130,180,170), QtGui.QColor(240,128,128,170),
            QtGui.QColor(60,179,113,170), QtGui.QColor(186,85,211,170),
            QtGui.QColor(210,180,140,170), QtGui.QColor(255,165,0,170),
            QtGui.QColor(128,128,128,170)
        ]
        line_colors = [
            QtGui.QColor(31,119,180), QtGui.QColor(255,127,14),
            QtGui.QColor(44,160,44), QtGui.QColor(148,103,189),
            QtGui.QColor(140,86,75), QtGui.QColor(227,119,194),
            QtGui.QColor(90,90,90)
        ]

        group_w = max(6.0, min(36.0, pw / max(20, len(plot_rows)*1.4)))
        single_w = max(2.0, group_w / max(1, len(active_daily)))
        for r in plot_rows:
            x_center = xp(r["DateTime"])
            for idx, n in enumerate(active_daily):
                v = self.f(r.get(n, 0.0))
                if v <= 0:
                    continue
                x = x_center - group_w / 2 + idx * single_w
                y0 = yd(0.0)
                y1 = yd(v)
                rect = QtCore.QRectF(x, y1, max(1.0, single_w * 0.85), max(1.0, y0-y1))
                self.scene_obj.addRect(rect, QtGui.QPen(QtCore.Qt.PenStyle.NoPen), QtGui.QBrush(bar_colors[idx % len(bar_colors)]))

        for idx, n in enumerate(active_cum):
            vals = cum_values.get(n, {})
            pen = QtGui.QPen(line_colors[idx % len(line_colors)], 2)
            prev = None
            for r in plot_rows:
                x = r.get("DateTime")
                if x not in vals:
                    prev = None
                    continue
                pt = QtCore.QPointF(xp(x), yc(vals[x]))
                self.scene_obj.addEllipse(pt.x()-2.4, pt.y()-2.4, 4.8, 4.8, pen, QtGui.QBrush(line_colors[idx % len(line_colors)]))
                if prev is not None:
                    self.scene_obj.addLine(prev.x(), prev.y(), pt.x(), pt.y(), pen)
                prev = pt

        lx = left
        ly = h - 34
        for idx, n in enumerate(active_daily[:4] + active_cum[:4]):
            if lx > w - 120:
                break
            color = bar_colors[idx % len(bar_colors)] if n in active_daily else line_colors[idx % len(line_colors)]
            self.scene_obj.addRect(lx, ly-6, 10, 10, QtGui.QPen(QtCore.Qt.PenStyle.NoPen), QtGui.QBrush(color))
            self.txt(lx+14, ly, n, 8, False, False)
            lx += min(140, 24 + len(n)*7)

        title = "WaterSystem Summary Chart  (rows=%d, visible=%d, plotted=%d)" % (len(self.summary_rows), len(visible_rows), len(plot_rows))
        if self.manual_x0 and self.manual_x1:
            title += "  [zoom: double-click to reset]"
        if daily_max <= 1.15 and cum_max <= 1.15:
            title += "  [all plotted values are zero/near zero]"
        self.txt(left+pw/2, 18, title, 13, True, True)
        self.last_status = title

    def set_tooltip_mode(self, mode):
        self.tooltip_mode = mode or "Point"
        self.hide_cursor_tooltip()

    def hide_cursor_tooltip(self):
        try:
            for attr in ["tooltip_item", "tooltip_bg", "crosshair_v", "crosshair_h"]:
                item = getattr(self, attr, None)
                if item:
                    self.scene_obj.removeItem(item)
                    setattr(self, attr, None)
        except Exception:
            self.tooltip_item = None
            self.tooltip_bg = None
            self.crosshair_v = None
            self.crosshair_h = None

    def draw_crosshair(self, scene_pos):
        try:
            for attr in ["crosshair_v", "crosshair_h"]:
                item = getattr(self, attr, None)
                if item:
                    self.scene_obj.removeItem(item)
                    setattr(self, attr, None)
            pen = QtGui.QPen(QtGui.QColor(80, 80, 80, 150), 1.0, QtCore.Qt.PenStyle.DashLine)
            self.crosshair_v = self.scene_obj.addLine(scene_pos.x(), self.plot_top, scene_pos.x(), self.plot_top + self.plot_ph, pen)
            self.crosshair_h = self.scene_obj.addLine(self.plot_left, scene_pos.y(), self.plot_left + self.plot_pw, scene_pos.y(), pen)
            self.crosshair_v.setZValue(9996)
            self.crosshair_h.setZValue(9996)
        except Exception:
            pass

    def show_cursor_tooltip(self, scene_pos):
        try:
            if self.tooltip_mode == "Off":
                self.hide_cursor_tooltip()
                return
            if self.drag_start is not None or self.pan_start is not None or not self.in_plot_area(scene_pos) or not self.last_plot_rows:
                self.hide_cursor_tooltip()
                return
            self.draw_crosshair(scene_pos)

            target_dt = self.scene_x_to_datetime(scene_pos.x())
            if not target_dt:
                self.hide_cursor_tooltip()
                return
            nearest = min(self.last_plot_rows, key=lambda r: abs((r.get("DateTime") - target_dt).total_seconds()) if isinstance(r.get("DateTime"), dt.datetime) else 10**18)
            d = nearest.get("DateTime")
            if not isinstance(d, dt.datetime):
                self.hide_cursor_tooltip()
                return

            if self.tooltip_mode == "Crosshair":
                lines = [d.strftime("%Y/%m/%d %H:%M")]
                for n in self.last_active_daily:
                    lines.append("%s: %.2f" % (n, self.f(nearest.get(n, 0.0))))
                for n in self.last_active_cum:
                    vals = self.last_cum_values.get(n, {})
                    v = vals.get(d, None)
                    if v is not None:
                        lines.append("%s: %.2f" % (n, v))
                txt = "\n".join(lines[:20])
            else:
                candidates = []
                for n in self.last_active_daily:
                    v = self.f(nearest.get(n, 0.0))
                    if v != 0:
                        candidates.append((n, v))
                for n in self.last_active_cum:
                    vals = self.last_cum_values.get(n, {})
                    v = vals.get(d, None)
                    if v is not None:
                        candidates.append((n, v))
                if candidates:
                    n, v = candidates[0]
                    txt = "%s\n%s: %.2f" % (d.strftime("%Y/%m/%d %H:%M"), n, v)
                else:
                    txt = "%s\n(no visible value)" % d.strftime("%Y/%m/%d %H:%M")

            if self.tooltip_item:
                self.scene_obj.removeItem(self.tooltip_item)
                self.tooltip_item = None
            if self.tooltip_bg:
                self.scene_obj.removeItem(self.tooltip_bg)
                self.tooltip_bg = None

            self.tooltip_item = self.scene_obj.addText(txt)
            font = QtGui.QFont(); font.setPointSize(9)
            self.tooltip_item.setFont(font)
            self.tooltip_item.setDefaultTextColor(QtGui.QColor(30, 30, 30))
            br = self.tooltip_item.boundingRect()
            x, y = scene_pos.x() + 14, scene_pos.y() + 14
            scene_rect = self.scene_obj.sceneRect()
            if x + br.width() + 16 > scene_rect.right():
                x = scene_pos.x() - br.width() - 24
            if y + br.height() + 16 > scene_rect.bottom():
                y = scene_pos.y() - br.height() - 24
            x, y = max(6, x), max(6, y)
            self.tooltip_bg = self.scene_obj.addRect(QtCore.QRectF(x - 6, y - 4, br.width() + 12, br.height() + 8), QtGui.QPen(QtGui.QColor(110,110,110)), QtGui.QBrush(QtGui.QColor(255,255,230,235)))
            self.tooltip_bg.setZValue(9998)
            self.tooltip_item.setZValue(9999)
            self.tooltip_item.setPos(x, y)
        except Exception:
            self.hide_cursor_tooltip()

    def in_plot_area(self, p):
        return (self.plot_left <= p.x() <= self.plot_left + self.plot_pw and
                self.plot_top <= p.y() <= self.plot_top + self.plot_ph)

    def scene_x_to_datetime(self, sx):
        if not self.current_x0 or not self.current_x1:
            return None
        ratio = (sx - self.plot_left) / float(max(1, self.plot_pw))
        ratio = max(0.0, min(1.0, ratio))
        span = (self.current_x1 - self.current_x0).total_seconds()
        return self.current_x0 + dt.timedelta(seconds=span * ratio)

    def mouseDoubleClickEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self.hide_cursor_tooltip()
            self.manual_x0 = None
            self.manual_x1 = None
            self.redraw()
            event.accept()
            return
        super().mouseDoubleClickEvent(event)

    def mousePressEvent(self, event):
        p = self.mapToScene(event.position().toPoint())
        self.hide_cursor_tooltip()
        if event.button() == QtCore.Qt.MouseButton.LeftButton and self.in_plot_area(p):
            self.drag_start = p
            if self.drag_item:
                self.scene_obj.removeItem(self.drag_item)
                self.drag_item = None
            event.accept()
            return
        if event.button() == QtCore.Qt.MouseButton.RightButton and self.in_plot_area(p):
            self.pan_start = p
            self.pan_x0 = self.current_x0
            self.pan_x1 = self.current_x1
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        p = self.mapToScene(event.position().toPoint())
        if self.drag_start is not None:
            x0 = max(self.plot_left, min(self.drag_start.x(), p.x()))
            x1 = min(self.plot_left + self.plot_pw, max(self.drag_start.x(), p.x()))
            y0 = self.plot_top
            rect = QtCore.QRectF(x0, y0, max(1, x1-x0), self.plot_ph)
            if self.drag_item:
                self.scene_obj.removeItem(self.drag_item)
            pen = QtGui.QPen(QtGui.QColor(30, 120, 220), 1.5, QtCore.Qt.PenStyle.DashLine)
            brush = QtGui.QBrush(QtGui.QColor(30, 120, 220, 35))
            self.drag_item = self.scene_obj.addRect(rect, pen, brush)
            event.accept()
            return
        if self.pan_start is not None and self.pan_x0 and self.pan_x1:
            dx = p.x() - self.pan_start.x()
            span_sec = (self.pan_x1 - self.pan_x0).total_seconds()
            shift_sec = -dx / float(max(1, self.plot_pw)) * span_sec
            self.manual_x0 = self.pan_x0 + dt.timedelta(seconds=shift_sec)
            self.manual_x1 = self.pan_x1 + dt.timedelta(seconds=shift_sec)
            self.redraw()
            event.accept()
            return
        if self.drag_start is None and self.pan_start is None:
            self.show_cursor_tooltip(p)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton and self.drag_start is not None:
            p = self.mapToScene(event.position().toPoint())
            x0 = max(self.plot_left, min(self.drag_start.x(), p.x()))
            x1 = min(self.plot_left + self.plot_pw, max(self.drag_start.x(), p.x()))
            if self.drag_item:
                self.scene_obj.removeItem(self.drag_item)
                self.drag_item = None
            self.drag_start = None
            if abs(x1 - x0) > 10:
                d0 = self.scene_x_to_datetime(x0)
                d1 = self.scene_x_to_datetime(x1)
                if d0 and d1 and d1 > d0:
                    self.manual_x0 = d0
                    self.manual_x1 = d1
                    self.redraw()
            event.accept()
            return
        if event.button() == QtCore.Qt.MouseButton.RightButton and self.pan_start is not None:
            self.pan_start = None
            self.pan_x0 = None
            self.pan_x1 = None
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def wheelEvent(self, event):
        p = self.mapToScene(event.position().toPoint())
        if not self.in_plot_area(p) or not self.current_x0 or not self.current_x1:
            super().wheelEvent(event)
            return
        center = self.scene_x_to_datetime(p.x())
        if not center:
            super().wheelEvent(event)
            return

        factor = 0.80 if event.angleDelta().y() > 0 else 1.25
        span = max(60.0, (self.current_x1 - self.current_x0).total_seconds() * factor)
        left_sec = (center - self.current_x0).total_seconds()
        old_span = max(1.0, (self.current_x1 - self.current_x0).total_seconds())
        ratio = left_sec / old_span
        new_left = center - dt.timedelta(seconds=span * ratio)
        new_right = new_left + dt.timedelta(seconds=span)
        self.manual_x0 = new_left
        self.manual_x1 = new_right
        self.redraw()
        event.accept()


class AnalyzeWorker(QtCore.QThread):
    progress = QtCore.Signal(int, int, str, int, int, str, float)
    finished_ok = QtCore.Signal(list, list, int, int)
    failed = QtCore.Signal(str)

    def __init__(self, folder, result_file="", use_existing_result=False, parent=None):
        super(AnalyzeWorker, self).__init__(parent)
        self.folder = folder
        self.result_file = result_file
        self.use_existing_result = use_existing_result
        self.cancel_requested = False
        self.t0 = 0

    def cancel(self):
        self.cancel_requested = True

    def emit_progress(self, cur, total, name, added, skipped, stage):
        self.progress.emit(cur, total, name, added, skipped, stage, time.time() - self.t0)

    def run(self):
        try:
            self.t0 = time.time()

            base_file = resolve_resultsummary_base_file(self.folder, self.result_file, self.use_existing_result)
            self.emit_progress(0, 100, os.path.basename(base_file or "(no existing ResultSummary)"), 0, 0, "Reading ResultSummary")
            def existing_progress(cur, total, stage):
                self.progress.emit(cur, total, os.path.basename(base_file or ""), 0, 0, stage, time.time() - self.t0)
            rows, keys = read_existing_rows(base_file, progress=existing_progress, cancel=lambda: self.cancel_requested)

            if self.cancel_requested:
                self.finished_ok.emit(rows, build_summary_rows(rows), 0, 0)
                return

            def scan_progress(scanned, curdir):
                self.progress.emit(0, 0, curdir, 0, 0, "Searching files: %d scanned" % scanned, time.time() - self.t0)

            files = collect_files(self.folder, progress=scan_progress, cancel=lambda: self.cancel_requested)
            total = len(files)
            added = 0
            skipped = 0
            parse_fail = 0
            duplicate = 0
            errors = []
            self.debug_info = {
                "mode": resultsummary_mode_text(self.result_file, self.use_existing_result),
                "base_file": base_file,
                "scanned": getattr(collect_files, "last_scan", {}).get("scanned", 0),
                "matched": total,
                "first_files": files[:10],
                "parse_fail": 0,
                "duplicate": 0,
                "errors": [],
            }

            self.emit_progress(0, total, "Matched files: %d" % total, added, skipped, "File scan complete")

            for i, p in enumerate(files, 1):
                if self.cancel_requested:
                    break

                name = os.path.basename(p)
                self.emit_progress(i, total, name, added, skipped, "Reading file %d/%d" % (i, total))

                try:
                    dtv = build_file_datetime_text(p)
                    size = os.path.getsize(p)

                    if name_key(name) in keys or result_key(name, dtv) in keys:
                        skipped += 1
                        duplicate += 1
                        continue

                    h = file_crc32_hex(p)
                    ck = content_key(h, size)
                    if ck in keys:
                        skipped += 1
                        duplicate += 1
                        continue

                    parsed = parse_watersystem_log(p)
                    if parsed:
                        rows.append(parsed)
                        keys.add(content_key(parsed["File CRC32"], parsed["File Size"]))
                        keys.add(result_key(parsed["File Name"], parsed["Start Datetime (YYYY-MM-DD HH:MM)"]))
                        keys.add(name_key(parsed["File Name"]))
                        added += 1
                    else:
                        skipped += 1
                        parse_fail += 1
                        if len(errors) < 20:
                            errors.append("Parse returned no data: %s" % p)
                except Exception as e:
                    skipped += 1
                    parse_fail += 1
                    if len(errors) < 20:
                        errors.append("%s : %s" % (p, e))

                self.debug_info = {
                    "mode": resultsummary_mode_text(self.result_file, self.use_existing_result),
                    "base_file": base_file,
                    "scanned": getattr(collect_files, "last_scan", {}).get("scanned", 0),
                    "matched": total,
                    "first_files": files[:10],
                    "parse_fail": parse_fail,
                    "duplicate": duplicate,
                    "errors": errors,
                }
                self.emit_progress(i, total, name, added, skipped, "Analyzing file %d/%d" % (i, total))

            self.debug_info = {
                "mode": resultsummary_mode_text(self.result_file, self.use_existing_result),
                "base_file": base_file,
                "scanned": getattr(collect_files, "last_scan", {}).get("scanned", 0),
                "matched": total,
                "first_files": files[:10],
                "parse_fail": parse_fail,
                "duplicate": duplicate,
                "errors": errors,
            }
            AnalyzeWorker.last_debug = self.debug_info
            self.emit_progress(total, total, "", added, skipped, "Building summary / ParseFail=%d / Duplicate=%d" % (parse_fail, duplicate))
            rows.sort(key=lambda r: coerce_datetime(r.get("Start Datetime (YYYY-MM-DD HH:MM)")) or dt.datetime.min)
            summary = build_summary_rows(rows)
            self.finished_ok.emit(rows, summary, added, skipped)

        except Exception:
            self.failed.emit(traceback.format_exc())



# =====================================
# Version 7.3 Data Foundation
# =====================================
DATA_SCHEMA_VERSION = "1.0"
PROJECT_SCHEMA_VERSION = "1.0"

SECTION_MASTER_HEADERS = [
    "Section ID", "Hospital Name", "Serial Number", "Component Type",
    "Section Start", "Section End", "Section Type",
    "Replacement Type", "Failure Type", "Failure Type Version",
    "Data State", "Data Confidence", "Section Quality",
    "Use For Reliability", "Use For Forecast", "Censored",
    "Outlier", "Outlier Reason", "Manual Review",
    "Source", "Metadata JSON", "Comment", "Created At", "Updated At",
    "Schema Version", "Algorithm Version"
]
FAILURE_TYPE_MASTER_HEADERS = [
    "Component Type", "Failure Type", "Failure Code", "Description",
    "Active", "Use For Reliability", "Use For Forecast",
    "Default Severity", "Version", "Created At", "Updated At", "Comment"
]
REPLACEMENT_TYPE_MASTER_HEADERS = [
    "Replacement Type", "Description", "Use For Reliability", "Use For Forecast",
    "Censored", "Failure Event", "Default Section End", "Active",
    "Version", "Created At", "Updated At", "Comment"
]
RELIABILITY_MODEL_CONFIG_HEADERS = [
    "Model Name", "Model Type", "Active", "Component Type",
    "Failure Type Scope", "Use Censored Data", "Minimum Section Quality",
    "Minimum Samples", "Default Weight", "Plugin Name",
    "Model Version", "Config JSON", "Comment"
]
PROJECT_MANIFEST_HEADERS = ["Key", "Value"]


def df_now():
    return dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")


def df_master_path(filename):
    try:
        return master_path(filename)
    except Exception:
        root = ensure_data_folders()
        master = os.path.join(root, "Master")
        os.makedirs(master, exist_ok=True)
        return os.path.join(master, filename)


def df_project_path():
    path = os.path.join(ensure_data_folders(), "Project")
    os.makedirs(path, exist_ok=True)
    return path


def df_docs_path():
    path = os.path.join(df_project_path(), "docs")
    os.makedirs(path, exist_ok=True)
    return path


def df_write_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def df_ensure_workbook(path, headers, sheet_name):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(c)].width = max(14, min(36, len(h) + 4))
        ws.freeze_panes = "A2"
        wb.save(path)
    else:
        wb = load_workbook(path)
        ws = wb.active
        current = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
        changed = False
        for h in headers:
            if h not in current:
                ws.cell(1, ws.max_column + 1, h)
                changed = True
        if changed:
            wb.save(path)


def df_append_defaults(path, headers, rows):
    wb = load_workbook(path)
    ws = wb.active
    if ws.max_row <= 1:
        for row in rows:
            r = ws.max_row + 1
            for c, h in enumerate(headers, 1):
                ws.cell(r, c, row.get(h, ""))
        wb.save(path)


def ensure_project_metadata_files():
    project = df_project_path()
    docs = df_docs_path()
    now = df_now()

    manifest = {
        "project_name": "WaterSystem Reliability Project",
        "project_id": str(uuid.uuid4()),
        "created_at": now,
        "last_modified": now,
        "software_version": APP_VERSION,
        "schema_version": DATA_SCHEMA_VERSION,
        "project_schema_version": PROJECT_SCHEMA_VERSION,
        "platform": "Medical Reliability Platform",
        "primary_component": "DO Module",
        "status": "Data Foundation initialized"
    }
    files = {
        "manifest.json": manifest,
        "metadata.json": {
            "metadata_version": "1.0",
            "data_policy": "Raw data never changes",
            "reproducibility_policy": "Analysis requires metadata, engine version, schema version, and inputs."
        },
        "engine.json": {
            "runtime_engine": ALGORITHM_VERSION,
            "reliability_engine": ALGORITHM_VERSION,
            "forecast_engine": "Not calculated yet",
            "supported_models": ["Standard Statistics", "Weibull", "Kaplan-Meier", "Nelson-Aalen", "MTBF", "MTTF", "Hazard Rate"]
        },
        "schema.json": {
            "schema_version": DATA_SCHEMA_VERSION,
            "section_master": SECTION_MASTER_HEADERS,
            "failure_type_master": FAILURE_TYPE_MASTER_HEADERS,
            "replacement_type_master": REPLACEMENT_TYPE_MASTER_HEADERS,
            "reliability_model_config": RELIABILITY_MODEL_CONFIG_HEADERS
        },
        "settings.json": {
            "mode": "User",
            "maintenance_mode_planned": True,
            "snapshot_before_migration": True,
            "module_update_dry_run_required": True
        },
        "migration.json": {
            "migration_history": [],
            "current_schema_version": DATA_SCHEMA_VERSION,
            "rollback_policy": "Snapshot before migration"
        }
    }
    for name, data in files.items():
        path = os.path.join(project, name)
        if not os.path.exists(path):
            df_write_json(path, data)

    docs_data = {
        "Architecture.md": "# Architecture\n\nMedical Reliability Platform principles:\n\n- Everything is a Project\n- Raw Data Never Changes\n- Everything is Reproducible\n- Everything is Explainable\n- Everything is Upgradeable\n- Everything is Auditable\n- Backward Compatibility First\n",
        "Roadmap.md": "# Roadmap\n\nM1 Data Foundation / Version 7.3\nM2 Reliability Models / Version 7.4\nM3 Forecast Engine / Version 7.5\nM4 Learning Engine / Version 7.6\nM5 Platform Foundation / Version 8.0\n",
        "DecisionLog.md": "# Decision Log\n\n## Decision 001\nScheduled Replacement is treated as censored data by default because the failure time is not observed.\n\n## Decision 002\nFailure Type is defined per Component Type.\n\n## Decision 003\nRaw Log must not be overwritten.\n",
        "DataModel.md": "# Data Model\n\nProject -> Hospital Name -> Serial Number -> Component Type -> Section -> Runtime/Reliability/Forecast\n\nSection is the main analysis unit.\n",
        "ReleaseNotes.md": "# Release Notes\n\n## Version 7.3 DataFoundation\n- Section Master\n- Failure Type Master\n- Replacement Type Master\n- Reliability Model Config\n- Project JSON metadata and docs foundation\n"
    }
    for name, content in docs_data.items():
        path = os.path.join(docs, name)
        if not os.path.exists(path):
            Path(path).write_text(content, encoding="utf-8")


def ensure_data_foundation_files():
    section_path = df_master_path("Section_Master.xlsx")
    failure_path = df_master_path("Failure_Type_Master.xlsx")
    repl_path = df_master_path("Replacement_Type_Master.xlsx")
    model_path = df_master_path("Reliability_Model_Config.xlsx")
    manifest_path = df_master_path("Project_Manifest.xlsx")

    df_ensure_workbook(section_path, SECTION_MASTER_HEADERS, "Section_Master")
    df_ensure_workbook(failure_path, FAILURE_TYPE_MASTER_HEADERS, "Failure_Type_Master")
    df_ensure_workbook(repl_path, REPLACEMENT_TYPE_MASTER_HEADERS, "Replacement_Type_Master")
    df_ensure_workbook(model_path, RELIABILITY_MODEL_CONFIG_HEADERS, "Reliability_Model_Config")
    df_ensure_workbook(manifest_path, PROJECT_MANIFEST_HEADERS, "Project_Manifest")

    now = df_now()
    df_append_defaults(failure_path, FAILURE_TYPE_MASTER_HEADERS, [
        {"Component Type":"DO Module","Failure Type":"Water in Vacuum Cup","Failure Code":"DO-WATER-CUP","Description":"Water accumulation in vacuum cup","Active":"Yes","Use For Reliability":"Yes","Use For Forecast":"Yes","Default Severity":"Medium","Version":"1.0","Created At":now,"Updated At":now},
        {"Component Type":"DO Module","Failure Type":"Slow Degas","Failure Code":"DO-SLOW-DEGAS","Description":"Degas is slower than expected","Active":"Yes","Use For Reliability":"Yes","Use For Forecast":"Yes","Default Severity":"Medium","Version":"1.0","Created At":now,"Updated At":now},
        {"Component Type":"DO Module","Failure Type":"Degas Minimum High","Failure Code":"DO-MIN-HIGH","Description":"Degas minimum value remains high","Active":"Yes","Use For Reliability":"Yes","Use For Forecast":"Yes","Default Severity":"High","Version":"1.0","Created At":now,"Updated At":now},
        {"Component Type":"DO Module","Failure Type":"Unknown","Failure Code":"UNKNOWN","Description":"Unknown failure type","Active":"Yes","Use For Reliability":"Conditional","Use For Forecast":"Conditional","Default Severity":"Unknown","Version":"1.0","Created At":now,"Updated At":now}
    ])

    df_append_defaults(repl_path, REPLACEMENT_TYPE_MASTER_HEADERS, [
        {"Replacement Type":"Failure","Description":"Replacement caused by failure","Use For Reliability":"Yes","Use For Forecast":"Yes","Censored":"No","Failure Event":"Yes","Default Section End":"Yes","Active":"Yes","Version":"1.0","Created At":now,"Updated At":now},
        {"Replacement Type":"Scheduled Replacement","Description":"Planned replacement; failure time not observed","Use For Reliability":"Yes","Use For Forecast":"Conditional","Censored":"Yes","Failure Event":"No","Default Section End":"Yes","Active":"Yes","Version":"1.0","Created At":now,"Updated At":now},
        {"Replacement Type":"Hospital Request","Description":"Hospital requested replacement; not failure-related by default","Use For Reliability":"No","Use For Forecast":"No","Censored":"No","Failure Event":"No","Default Section End":"Yes","Active":"Yes","Version":"1.0","Created At":now,"Updated At":now},
        {"Replacement Type":"Company Instruction","Description":"Company/service instruction replacement","Use For Reliability":"No","Use For Forecast":"No","Censored":"No","Failure Event":"No","Default Section End":"Yes","Active":"Yes","Version":"1.0","Created At":now,"Updated At":now},
        {"Replacement Type":"Unknown","Description":"Unknown replacement type","Use For Reliability":"Conditional","Use For Forecast":"Conditional","Censored":"Conditional","Failure Event":"Unknown","Default Section End":"Yes","Active":"Yes","Version":"1.0","Created At":now,"Updated At":now}
    ])

    df_append_defaults(model_path, RELIABILITY_MODEL_CONFIG_HEADERS, [
        {"Model Name":"Standard Statistics","Model Type":"Statistics","Active":"Yes","Component Type":"All","Failure Type Scope":"All","Use Censored Data":"No","Minimum Section Quality":50,"Minimum Samples":3,"Default Weight":1.0,"Plugin Name":"standard_statistics","Model Version":"1.0","Config JSON":"{\"method\":\"mean_median_percentile\"}"},
        {"Model Name":"Weibull","Model Type":"Parametric","Active":"Planned","Component Type":"All","Failure Type Scope":"Failure only","Use Censored Data":"Yes","Minimum Section Quality":70,"Minimum Samples":5,"Default Weight":1.0,"Plugin Name":"weibull","Model Version":"0.1-planned","Config JSON":"{\"distribution\":\"weibull\",\"censored\":true}"},
        {"Model Name":"Kaplan-Meier","Model Type":"Survival","Active":"Planned","Component Type":"All","Failure Type Scope":"All","Use Censored Data":"Yes","Minimum Section Quality":70,"Minimum Samples":5,"Default Weight":1.0,"Plugin Name":"kaplan_meier","Model Version":"0.1-planned","Config JSON":"{\"survival_model\":\"kaplan_meier\"}"},
        {"Model Name":"Nelson-Aalen","Model Type":"Survival","Active":"Planned","Component Type":"All","Failure Type Scope":"All","Use Censored Data":"Yes","Minimum Section Quality":70,"Minimum Samples":5,"Default Weight":1.0,"Plugin Name":"nelson_aalen","Model Version":"0.1-planned","Config JSON":"{\"hazard_model\":\"nelson_aalen\"}"},
        {"Model Name":"MTBF","Model Type":"Metric","Active":"Planned","Component Type":"All","Failure Type Scope":"Failure only","Use Censored Data":"No","Minimum Section Quality":60,"Minimum Samples":3,"Default Weight":1.0,"Plugin Name":"mtbf","Model Version":"0.1-planned","Config JSON":"{\"metric\":\"mtbf\"}"}
    ])

    ensure_project_metadata_files()
    return {
        "Section_Master.xlsx": section_path,
        "Failure_Type_Master.xlsx": failure_path,
        "Replacement_Type_Master.xlsx": repl_path,
        "Reliability_Model_Config.xlsx": model_path,
        "Project_Manifest.xlsx": manifest_path,
        "Project Folder": df_project_path()
    }


def build_data_foundation_report():
    paths = ensure_data_foundation_files()
    lines = ["Data Foundation initialized.", "", "Schema Version: %s" % DATA_SCHEMA_VERSION, "", "Created/checked:"]
    for k, v in paths.items():
        lines.append("- %s: %s" % (k, v))
    return "\n".join(lines), paths


class DataFoundationDialog(QtWidgets.QDialog):
    def __init__(self, report_text, paths, parent=None):
        super(DataFoundationDialog, self).__init__(parent)
        self.setWindowTitle("Data Foundation")
        self.resize(900, 600)
        self.report_text = report_text
        self.paths = paths or {}
        self._ui()

    def _ui(self):
        layout = QVBoxLayout(self)
        label = QLabel("<b>Version 7.3 Data Foundation</b><br>Project metadata, Section Master, Failure Type Master, Replacement Type Master, and Model Config are initialized.")
        label.setWordWrap(True)
        layout.addWidget(label)
        t = QtWidgets.QTextEdit()
        t.setReadOnly(True)
        t.setPlainText(self.report_text)
        layout.addWidget(t, 1)
        buttons = QHBoxLayout()
        b = QPushButton("Open Master Folder")
        b.clicked.connect(lambda: QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(os.path.dirname(df_master_path("Section_Master.xlsx")))))
        buttons.addWidget(b)
        b = QPushButton("Open Project Folder")
        b.clicked.connect(lambda: QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(df_project_path())))
        buttons.addWidget(b)
        b = QPushButton("Close")
        b.clicked.connect(self.accept)
        buttons.addWidget(b)
        layout.addLayout(buttons)


# =====================================
# Version 7.3a Dashboard Design Preview
# =====================================

def dashboard_safe_float(v, default=0.0):
    try:
        if v in ("", None):
            return default
        return float(v)
    except Exception:
        return default


def dashboard_latest_xlsx_rows(filename, max_rows=50):
    try:
        path = data_foundation_master_path(filename) if "data_foundation_master_path" in globals() else master_path(filename)
    except Exception:
        return []
    rows = []
    try:
        if not os.path.exists(path):
            return rows
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for i, h in enumerate(headers):
                if h:
                    row[h] = values[i] if i < len(values) else ""
            rows.append(row)
        try:
            wb.close()
        except Exception:
            pass
    except Exception:
        return []
    return rows[-max_rows:]


def dashboard_context(summary_rows):
    explain = dashboard_latest_xlsx_rows("Explainable_Reliability.xlsx", 50)
    trend = dashboard_latest_xlsx_rows("Reliability_Trend.xlsx", 50)
    latest = explain[-1] if explain else {}
    if not latest:
        has_data = bool(summary_rows)
        latest = {
            "Forecast Readiness": "READY WITH CAUTION" if has_data else "NOT READY",
            "Overall Confidence Score": 75 if has_data else 0,
            "Overall Confidence Stars": "★★★☆☆" if has_data else "★☆☆☆☆",
            "Coverage Score": 80 if has_data else 0,
            "Data Continuity Score": 75 if has_data else 0,
            "Gap Quality Score": 75 if has_data else 0,
            "Runtime Confidence Score": 80 if has_data else 0,
            "Replacement Quality Score": 50,
            "Prediction Stability Score": 50,
            "Data Quality Score": 90 if has_data else 0,
            "Overall Explanation": "Dashboard preview fallback. Run Runtime Reconstruction / Explainable Reliability for calculated values."
        }
    return latest, trend, explain


class DashboardCard(QtWidgets.QFrame):
    def __init__(self, title, value, subtitle="", parent=None):
        super(DashboardCard, self).__init__(parent)
        self.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.setMinimumHeight(110)
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("<b>%s</b>" % title))
        lay.addWidget(QLabel("<span style='font-size:28px; font-weight:bold;'>%s</span>" % value))
        sub = QLabel("<span style='color:#666;'>%s</span>" % subtitle)
        sub.setWordWrap(True)
        lay.addWidget(sub)


class MiniTrendWidget(QtWidgets.QWidget):
    def __init__(self, trend_rows, parent=None):
        super(MiniTrendWidget, self).__init__(parent)
        self.trend_rows = trend_rows or []
        self.setMinimumHeight(220)

    def paintEvent(self, event):
        p = QtGui.QPainter(self)
        p.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        r = self.rect().adjusted(14,14,-14,-14)
        f = QtGui.QFont(); f.setPointSize(12); f.setBold(True); p.setFont(f)
        p.setPen(QtGui.QColor(35,35,35)); p.drawText(r.left(), r.top(), "Reliability Trend")
        plot = QtCore.QRectF(r.left()+30, r.top()+40, r.width()-60, r.height()-70)
        p.setPen(QtGui.QPen(QtGui.QColor(225,225,225), 1))
        for i in range(5):
            y = plot.top() + plot.height()*i/4
            p.drawLine(plot.left(), y, plot.right(), y)
        vals = []
        for row in self.trend_rows[-12:]:
            vals.append(dashboard_safe_float(row.get("Overall Confidence Score"), 0))
        if not vals:
            vals = [70, 73, 75, 78, 80, 82]
        n = max(1, len(vals)-1)
        pts = []
        for i,v in enumerate(vals):
            pts.append(QtCore.QPointF(plot.left()+plot.width()*(i/n if n else 0), plot.bottom()-plot.height()*max(0,min(100,v))/100.0))
        p.setPen(QtGui.QPen(QtGui.QColor(70,145,220), 3))
        for a,b in zip(pts, pts[1:]):
            p.drawLine(a,b)
        p.setBrush(QtGui.QColor(70,145,220))
        for pt in pts:
            p.drawEllipse(pt, 4, 4)


class ScoreBarsWidget(QtWidgets.QWidget):
    def __init__(self, items, parent=None):
        super(ScoreBarsWidget, self).__init__(parent)
        self.items = items or []
        self.setMinimumHeight(280)

    def paintEvent(self, event):
        p = QtGui.QPainter(self); p.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        r = self.rect().adjusted(14,14,-14,-14)
        f = QtGui.QFont(); f.setPointSize(12); f.setBold(True); p.setFont(f)
        p.setPen(QtGui.QColor(35,35,35)); p.drawText(r.left(), r.top(), "Confidence Breakdown")
        y = r.top()+42
        for label, score in self.items:
            score = dashboard_safe_float(score, 0)
            p.setPen(QtGui.QColor(60,60,60)); f.setPointSize(9); f.setBold(True); p.setFont(f)
            p.drawText(r.left(), y+13, label)
            bar = QtCore.QRectF(r.left()+155, y, r.width()-230, 16)
            p.setPen(QtCore.Qt.PenStyle.NoPen); p.setBrush(QtGui.QColor(232,232,232)); p.drawRoundedRect(bar,8,8)
            color = QtGui.QColor(75,170,95) if score >= 85 else (QtGui.QColor(230,180,70) if score >= 65 else QtGui.QColor(220,90,80))
            fill = QtCore.QRectF(bar.left(), bar.top(), bar.width()*max(0,min(100,score))/100.0, bar.height())
            p.setBrush(color); p.drawRoundedRect(fill,8,8)
            p.setPen(QtGui.QColor(40,40,40)); f.setBold(False); p.setFont(f); p.drawText(bar.right()+10, y+13, "%.0f" % score)
            y += 34


class FailureMixWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(FailureMixWidget, self).__init__(parent)
        self.setMinimumHeight(220)
        self.items = [("Water in Vacuum Cup",35),("Slow Degas",28),("Degas Minimum High",20),("Unknown",17)]

    def paintEvent(self, event):
        p = QtGui.QPainter(self); p.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        r = self.rect().adjusted(14,14,-14,-14)
        f = QtGui.QFont(); f.setPointSize(12); f.setBold(True); p.setFont(f)
        p.setPen(QtGui.QColor(35,35,35)); p.drawText(r.left(), r.top(), "Failure Distribution")
        pie = QtCore.QRectF(r.left()+20, r.top()+45, 130, 130)
        total = sum(v for _,v in self.items); start=0
        colors=[QtGui.QColor(70,145,220), QtGui.QColor(245,170,70), QtGui.QColor(220,90,80), QtGui.QColor(160,160,160)]
        p.setPen(QtCore.Qt.PenStyle.NoPen)
        for i,(name,val) in enumerate(self.items):
            span=int(360*16*val/total); p.setBrush(colors[i]); p.drawPie(pie, start, span); start += span
        y=r.top()+48; f.setPointSize(9); f.setBold(False); p.setFont(f)
        for i,(name,val) in enumerate(self.items):
            p.setBrush(colors[i]); p.drawRect(r.left()+180,y-9,10,10)
            p.setPen(QtGui.QColor(60,60,60)); p.drawText(r.left()+198, y, "%s  %d%%" % (name,val)); y += 28


class ForecastTimelineWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(ForecastTimelineWidget, self).__init__(parent)
        self.setMinimumHeight(180)

    def paintEvent(self, event):
        p=QtGui.QPainter(self); p.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        r=self.rect().adjusted(14,14,-14,-14)
        f=QtGui.QFont(); f.setPointSize(12); f.setBold(True); p.setFont(f)
        p.setPen(QtGui.QColor(35,35,35)); p.drawText(r.left(), r.top(), "Forecast Timeline")
        y=r.top()+82; x1=r.left()+50; x2=r.right()-50
        p.setPen(QtGui.QPen(QtGui.QColor(80,80,80),2)); p.drawLine(x1,y,x2,y)
        p.setBrush(QtGui.QColor(70,145,220)); p.drawEllipse(QtCore.QPointF(x1,y),6,6)
        p.setBrush(QtGui.QColor(230,180,70)); p.drawRoundedRect(QtCore.QRectF(x2-120,y-10,95,20),8,8)
        p.setPen(QtGui.QColor(60,60,60)); f.setPointSize(9); f.setBold(False); p.setFont(f)
        p.drawText(x1-20,y+28,"Today"); p.drawText(x2-140,y+28,"Forecast Window"); p.drawText(x2-35,y+28,"Replace")


class DashboardPreviewDialog(QtWidgets.QDialog):
    def __init__(self, latest, trend_rows, explain_rows, parent=None):
        super(DashboardPreviewDialog, self).__init__(parent)
        self.setWindowTitle("Medical Reliability Dashboard Preview")
        self.resize(1280, 820)
        self.latest = latest or {}; self.trend_rows = trend_rows or []; self.explain_rows = explain_rows or []
        self._ui()

    def _ui(self):
        layout=QVBoxLayout(self)
        scroll=QScrollArea(); scroll.setWidgetResizable(True)
        body=QWidget(); bl=QVBoxLayout(body); scroll.setWidget(body); layout.addWidget(scroll,1)
        readiness=self.latest.get("Forecast Readiness","READY WITH CAUTION")
        overall=dashboard_safe_float(self.latest.get("Overall Confidence Score"),75)
        stars=self.latest.get("Overall Confidence Stars","★★★☆☆")
        if readiness == "READY":
            title, color = "HEALTHY / READY FOR FORECAST", "#2e9d55"
        elif readiness == "READY WITH CAUTION":
            title, color = "MONITOR / FORECAST WITH CAUTION", "#c99722"
        else:
            title, color = "INSUFFICIENT DATA", "#c94c44"
        header=QLabel("<div style='font-size:28px; font-weight:bold;'>Medical Reliability Dashboard</div><div style='font-size:22px; font-weight:bold; color:%s;'>%s</div><div style='font-size:18px;'>Reliability: <b>%.1f</b> %s</div>" % (color,title,overall,stars))
        header.setWordWrap(True); header.setMinimumHeight(110); bl.addWidget(header)
        cards=QHBoxLayout()
        cards.addWidget(DashboardCard("Health Score","%.1f"%overall,"Overall reliability confidence"))
        cards.addWidget(DashboardCard("Remaining Life","Preview","Forecast Engine starts in Version 7.5"))
        cards.addWidget(DashboardCard("Confidence",stars,"Explainable reliability"))
        cards.addWidget(DashboardCard("Recommendation","Monitor","Prototype rule-based display"))
        bl.addLayout(cards)
        row1=QHBoxLayout(); row1.addWidget(MiniTrendWidget(self.trend_rows)); row1.addWidget(ForecastTimelineWidget()); bl.addLayout(row1)
        row2=QHBoxLayout(); row2.addWidget(FailureMixWidget())
        breakdown=[("Coverage",self.latest.get("Coverage Score",0)),("Continuity",self.latest.get("Data Continuity Score",0)),("Gap quality",self.latest.get("Gap Quality Score",0)),("Runtime",self.latest.get("Runtime Confidence Score",0)),("Replacement",self.latest.get("Replacement Quality Score",0)),("Prediction",self.latest.get("Prediction Stability Score",0)),("Data quality",self.latest.get("Data Quality Score",0))]
        row2.addWidget(ScoreBarsWidget(breakdown)); bl.addLayout(row2)
        why=QLabel("<b>Why?</b><br>%s" % self.latest.get("Overall Explanation","Run Explainable Reliability to populate detailed reasons."))
        why.setWordWrap(True); why.setMinimumHeight(90); bl.addWidget(why)
        details=QtWidgets.QGroupBox("Details"); dl=QVBoxLayout(details); txt=QtWidgets.QTextEdit(); txt.setReadOnly(True); txt.setPlainText(json.dumps(self.latest, ensure_ascii=False, indent=2, default=str)); dl.addWidget(txt); bl.addWidget(details)
        b=QPushButton("Close"); b.clicked.connect(self.accept); layout.addWidget(b)

class AnalyzerWindow(QMainWindow):
    def __init__(self):
        super(AnalyzerWindow, self).__init__()
        self.setWindowTitle("%s %s" % (APP_NAME, APP_VERSION))
        self.resize(1280, 800)

        self.folder = ""
        self.result_file = ""
        self.rows = []
        self.summary_rows = []
        self.worker = None
        self.checks = {}

        self.resize(1600, 900)
        self.setMinimumSize(1100, 720)
        self.setStyleSheet("QPushButton{min-height:24px;} QLineEdit{min-height:24px;} QCheckBox{min-height:22px;}")
        self._ui()
        ensure_data_folders()
        self.status.setText("Ready. Data folder and master files are prepared.")

    def _ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        main = QHBoxLayout(root)

        panel = QWidget()
        panel.setMinimumWidth(260)
        panel.setMaximumWidth(420)
        panel.setFixedWidth(360)
        self.panel = panel
        panel.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Expanding)
        self.menu_collapsed = False
        self.menu_button_widgets = []
        pl = QVBoxLayout(panel)

        self.btn_menu_toggle = QPushButton("▼ Menu")
        self.btn_menu_toggle.clicked.connect(self.toggle_menu_panel)
        pl.addWidget(self.btn_menu_toggle)

        # Main workflow groups.
        self.project_group = self.create_group("Project", expanded=True)
        project_lay = self.project_group["layout"]
        pl.addWidget(self.project_group["box"])

        b = QPushButton("Select WaterSystem Folder")
        b.clicked.connect(self.select_folder)
        project_lay.addWidget(b)

        self.folder_box = QLineEdit()
        self.folder_box.setReadOnly(True)
        project_lay.addWidget(self.folder_box)

        b = QPushButton("Select ResultSummary File")
        b.clicked.connect(self.select_result_file)
        project_lay.addWidget(b)

        self.result_box = QLineEdit()
        self.result_box.setReadOnly(True)
        project_lay.addWidget(self.result_box)

        self.use_existing_result_check = QCheckBox("Use existing ResultSummary")
        self.use_existing_result_check.setChecked(False)
        self.use_existing_result_check.setToolTip("OFF: analyze only log files. ON: merge with latest ResultSummary if no explicit file is selected.")
        project_lay.addWidget(self.use_existing_result_check)

        self.analysis_group = self.create_group("Analysis", expanded=True)
        analysis_lay = self.analysis_group["layout"]
        pl.addWidget(self.analysis_group["box"])

        self.btn_analyze = QPushButton("Analyze / Update ResultSummary")
        self.btn_analyze.clicked.connect(self.start_analyze)
        analysis_lay.addWidget(self.btn_analyze)

        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.clicked.connect(self.cancel_analyze)
        self.btn_cancel.setEnabled(False)
        analysis_lay.addWidget(self.btn_cancel)

        b = QPushButton("Save / Update ResultSummary")
        b.clicked.connect(self.save_result_summary_action)
        analysis_lay.addWidget(b)

        b = QPushButton("Refresh Current Chart")
        b.clicked.connect(self.update_chart)
        analysis_lay.addWidget(b)

        self.forecast_group = self.create_group("Forecast", expanded=False)
        forecast_lay = self.forecast_group["layout"]
        pl.addWidget(self.forecast_group["box"])

        b = QPushButton("Calculate Forecast Now")
        b.clicked.connect(self.calculate_forecast_now)
        forecast_lay.addWidget(b)

        b = QPushButton("Forecast Foundation / Data Coverage")
        b.clicked.connect(self.forecast_foundation_action)
        forecast_lay.addWidget(b)

        b = QPushButton("Reliability Dashboard")
        b.clicked.connect(self.reliability_dashboard_action)
        forecast_lay.addWidget(b)

        b = QPushButton("Runtime Reconstruction")
        b.clicked.connect(self.runtime_reconstruction_action)
        forecast_lay.addWidget(b)

        b = QPushButton("Explainable Reliability")
        b.clicked.connect(self.runtime_reconstruction_action)
        forecast_lay.addWidget(b)

        b = QPushButton("Import Replacement History")
        b.clicked.connect(self.import_replacement_history_action)
        forecast_lay.addWidget(b)

        self.master_group = self.create_group("Masters", expanded=False)
        master_lay = self.master_group["layout"]
        pl.addWidget(self.master_group["box"])

        b = QPushButton("Master Foundation")
        b.clicked.connect(self.open_master_foundation)
        master_lay.addWidget(b)

        b = QPushButton("Data Foundation")
        b.clicked.connect(self.data_foundation_action)
        master_lay.addWidget(b)

        b = QPushButton("Dashboard Preview")
        b.clicked.connect(self.dashboard_preview_action)
        master_lay.addWidget(b)

        self.migration_group = self.create_group("Backup / Migration", expanded=False)
        migration_lay = self.migration_group["layout"]
        pl.addWidget(self.migration_group["box"])

        b = QPushButton("Export Release Package")
        b.clicked.connect(self.export_release_package)
        migration_lay.addWidget(b)

        b = QPushButton("Import Release Package")
        b.clicked.connect(self.import_release_package)
        migration_lay.addWidget(b)

        b = QPushButton("Project Backup")
        b.clicked.connect(self.project_backup)
        migration_lay.addWidget(b)

        b = QPushButton("Restore Project Backup")
        b.clicked.connect(self.restore_project_backup)
        migration_lay.addWidget(b)

        b = QPushButton("Migration Wizard")
        b.clicked.connect(self.migration_wizard)
        migration_lay.addWidget(b)

        b = QPushButton("Move Data Folder")
        b.clicked.connect(self.move_data_folder)
        migration_lay.addWidget(b)

        self.tools_group = self.create_group("Tools / Debug", expanded=False)
        tools_lay = self.tools_group["layout"]
        pl.addWidget(self.tools_group["box"])

        b = QPushButton("Scan Folder Check")
        b.clicked.connect(self.scan_folder_check)
        tools_lay.addWidget(b)

        b = QPushButton("Date Parse Check")
        b.clicked.connect(self.date_parse_check)
        tools_lay.addWidget(b)

        b = QPushButton("Show Analyze Debug")
        b.clicked.connect(self.show_analyze_debug)
        tools_lay.addWidget(b)

        b = QPushButton("Show Analyze Performance")
        b.clicked.connect(self.show_analyze_performance)
        tools_lay.addWidget(b)

        b = QPushButton("Show Current Data Count")
        b.clicked.connect(self.show_current_data_count)
        tools_lay.addWidget(b)

        b = QPushButton("ResultSummary Check")
        b.clicked.connect(self.resultsummary_check)
        tools_lay.addWidget(b)

        b = QPushButton("Validate Masters")
        b.clicked.connect(self.validate_masters)
        tools_lay.addWidget(b)

        b = QPushButton("Migration Check")
        b.clicked.connect(self.migration_check)
        tools_lay.addWidget(b)

        b = QPushButton("Backup Manager")
        b.clicked.connect(self.backup_manager)
        tools_lay.addWidget(b)

        b = QPushButton("Show Data Location")
        b.clicked.connect(self.show_data_location)
        tools_lay.addWidget(b)

        b = QPushButton("Open Data Folder")
        b.clicked.connect(self.open_data_folder)
        tools_lay.addWidget(b)

        self.progress = QProgressBar()
        pl.addWidget(self.progress)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        ctrl = QWidget()
        cl = QVBoxLayout(ctrl)
        scroll.setWidget(ctrl)
        pl.addWidget(scroll, 1)

        cl.addWidget(QLabel("<b>Chart Controls</b>"))

        self.tooltip_mode_combo = self.add_combo(cl, "Popup Mode", ["Point", "Crosshair", "Off"])
        self.tooltip_mode_combo.currentTextChanged.connect(lambda v: self.chart_view.set_tooltip_mode(v))

        self.all_check = QCheckBox("ALL")
        self.all_check.setChecked(True)
        self.all_check.setStyleSheet("font-weight: bold; text-decoration: underline;")
        self.all_check.stateChanged.connect(lambda _=None: self.set_all())
        cl.addWidget(self.all_check)

        self.daily_group_check = QCheckBox("DAILY")
        self.daily_group_check.setChecked(True)
        self.daily_group_check.setStyleSheet("font-weight: bold; text-decoration: underline;")
        self.daily_group_check.stateChanged.connect(lambda _=None: self.set_group("daily"))
        cl.addWidget(self.daily_group_check)
        for n in ["Treat", "Degas", "Clean", "Runtime"]:
            self.add_check(cl, n, "daily")

        self.cum_group_check = QCheckBox("CUMULATIVE")
        self.cum_group_check.setChecked(True)
        self.cum_group_check.setStyleSheet("font-weight: bold; text-decoration: underline;")
        self.cum_group_check.stateChanged.connect(lambda _=None: self.set_group("cum"))
        cl.addWidget(self.cum_group_check)
        for n in ["Treat Cum", "Degas Cum", "Clean Cum", "Runtime Cum"]:
            self.add_check(cl, n, "cum")

        self.pause_group_check = QCheckBox("PAUSE/ERROR")
        self.pause_group_check.setChecked(True)
        self.pause_group_check.setStyleSheet("font-weight: bold; text-decoration: underline;")
        self.pause_group_check.stateChanged.connect(lambda _=None: self.set_group("pause"))
        cl.addWidget(self.pause_group_check)
        for n in ["DEGAS_PAUSE", "TREAT_PAUSE", "ERROR", "DEGAS_PAUSE Cum", "TREAT_PAUSE Cum", "ERROR Cum"]:
            self.add_check(cl, n, "pause")

        # X Start / End are intentionally hidden. Chart uses drag zoom / double click reset.
        self.x_start = QComboBox()
        self.x_start.addItems(["Auto"])
        self.x_end = QComboBox()
        self.x_end.addItems(["Auto"])
        self.cum_start = self.add_combo(cl, "Cum Start Date / Time")
        self.cum_end = self.add_combo(cl, "Cum End Date / Time")
        self.display_size = self.add_combo(cl, "Display Size", ["Auto Fit", "Laptop", "Full HD", "QHD", "4K"])
        self.forecast_mode = self.add_combo(cl, "Forecast Mode", ["Failure only", "Preventive included", "Compare both"])
        self.forecast_scope = self.add_combo(cl, "Forecast Scope", ["Current SN", "Same Hospital", "Selected Hospitals", "All Hospitals"])
        self.missing_log_handling = self.add_combo(cl, "Missing Log Handling", ["Unknown", "Estimated Usage", "Conservative"])

        b = QPushButton("Clear change")
        b.clicked.connect(self.clear_change)
        cl.addWidget(b)
        cl.addStretch(1)

        self.status = QLabel("Select a folder. Folder selection does not start analysis.")
        self.status.setWordWrap(True)
        pl.addWidget(self.status)

        chart = QWidget()
        chl = QVBoxLayout(chart)
        self.chart_view = NativeChartView()
        self.chart_view.installEventFilter(self)
        try:
            self.chart_view.viewport().installEventFilter(self)
        except Exception:
            pass
        chl.addWidget(self.chart_view)

        main.addWidget(panel, 0)
        main.addWidget(chart, 1)


    def create_group(self, title, expanded=True):
        box = QtWidgets.QGroupBox()
        box.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Maximum)
        outer = QVBoxLayout(box)
        outer.setContentsMargins(6, 6, 6, 6)
        btn = QPushButton(("▼ " if expanded else "▶ ") + title)
        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setContentsMargins(8, 4, 4, 4)
        content.setVisible(expanded)
        outer.addWidget(btn)
        outer.addWidget(content)
        group = {"box": box, "button": btn, "content": content, "layout": lay, "title": title}
        def toggle():
            self.expand_only_group(group if not content.isVisible() else None)
        btn.clicked.connect(toggle)
        return group

    def expand_only_group(self, target_group):
        for grp_name in ["project_group", "analysis_group", "forecast_group", "master_group", "migration_group", "tools_group"]:
            grp = getattr(self, grp_name, None)
            if not grp:
                continue
            vis = (grp is target_group)
            grp["content"].setVisible(vis)
            grp["button"].setText(("▼ " if vis else "▶ ") + grp["title"])
        try:
            self.panel.setFixedWidth(360)
        except Exception:
            pass


    def collect_menu_button_widgets(self):
        try:
            widgets = []
            for grp_name in ["project_group", "analysis_group", "forecast_group", "master_group", "migration_group", "tools_group"]:
                grp = getattr(self, grp_name, None)
                if grp:
                    widgets.append(grp["box"])
            self.menu_button_widgets = widgets
        except Exception:
            self.menu_button_widgets = []


    def set_menu_collapsed(self, collapsed=True):
        self.menu_collapsed = collapsed
        self.collect_menu_button_widgets()
        for w in self.menu_button_widgets:
            w.setVisible(not collapsed)

        # Also hide path fields and result options when collapsed; keep checkboxes visible.
        for w in [getattr(self, "folder_box", None), getattr(self, "result_box", None), getattr(self, "use_existing_result_check", None)]:
            if w:
                w.setVisible(not collapsed)

        self.btn_menu_toggle.setText("▶ Menu" if collapsed else "▼ Menu")
        if collapsed:
            self.panel.setFixedWidth(260)
        else:
            self.panel.setFixedWidth(360)

    def toggle_menu_panel(self):
        self.set_menu_collapsed(not getattr(self, "menu_collapsed", False))

    def collapse_menu_after_chart_ready(self):
        self.set_menu_collapsed(True)

    def eventFilter(self, obj, event):
        try:
            if obj in (getattr(self, "chart_view", None), getattr(getattr(self, "chart_view", None), "viewport", lambda: None)()):
                if event.type() in (QtCore.QEvent.Type.MouseButtonPress, QtCore.QEvent.Type.Wheel):
                    self.set_menu_collapsed(True)
        except Exception:
            pass
        return super(AnalyzerWindow, self).eventFilter(obj, event)

    def add_check(self, layout, name, group=""):
        cb = QCheckBox("    " + name)
        cb.setChecked(True)
        cb.group = group
        cb.stateChanged.connect(lambda _=None: self.on_child_check_changed())
        self.checks[name] = cb
        layout.addWidget(cb)

    def add_combo(self, layout, label, values=None):
        layout.addWidget(QLabel(label))
        c = QComboBox()
        c.addItems(values or ["Auto"])
        c.currentIndexChanged.connect(self.update_chart)
        layout.addWidget(c)
        return c

    def apply_display_size(self):
        v = self.display_size.currentText()
        if v == "Laptop":
            self.resize(1180, 720)
        elif v == "Full HD":
            self.resize(1500, 900)
        elif v == "QHD":
            self.resize(1900, 1050)
        elif v == "4K":
            self.resize(2500, 1350)

    def select_folder(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "Select WaterSystem folder",
            self.folder or os.getcwd(),
            QtWidgets.QFileDialog.Option.ShowDirsOnly | QtWidgets.QFileDialog.Option.DontResolveSymlinks
        )
        if folder:
            self.folder = folder
            self.folder_box.setText(folder)
            self.status.setText("Folder selected. Click Analyze / Update ResultSummary.")

    def select_result_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Select ResultSummary file",
            self.folder or os.getcwd(),
            "Excel files (*.xlsx *.xlsm);;All files (*.*)"
        )
        if path:
            self.result_file = path
            self.result_box.setText(path)
            if not self.folder:
                self.folder = os.path.dirname(path)
                self.folder_box.setText(self.folder)
            try:
                self.show_progress_dialog("Updating", "Loading ResultSummary", "Reading workbook...", 0, allow_cancel=False)
                QApplication.processEvents()
                def rs_progress(cur, total, stage):
                    self.update_progress_dialog(stage, os.path.basename(path), cur, total, None)
                self.rows, _existing_keys = read_existing_rows(path, progress=rs_progress, cancel=None)

                self.update_progress_dialog("Loading ResultSummary", "Reading chart summary...", 0, 0)
                direct_summary = read_chart_summary_direct(path)

                self.update_progress_dialog("Loading ResultSummary", "Building chart data...", 0, 0)
                rebuilt_summary = build_summary_rows(self.rows)
                if rebuilt_summary:
                    self.summary_rows = rebuilt_summary
                elif direct_summary:
                    self.summary_rows = direct_summary
                else:
                    self.summary_rows = []

                self.update_progress_dialog("Loading ResultSummary", "Updating chart and controls...", 0, 0)
                self.refresh_dropdowns()
                self.update_chart()
                self.close_progress_dialog()
                self.status.setText("ResultSummary loaded. Result rows: %d / Chart rows: %d / Direct summary: %s" % (len(self.rows), len(self.summary_rows), "YES" if direct_summary else "NO"))
            except Exception:
                self.close_progress_dialog()
                self.status.setText("ResultSummary selected, but load failed.")
                QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def group_names(self, group):
        if group == "daily":
            return ["Treat", "Degas", "Clean", "Runtime"]
        if group == "cum":
            return ["Treat Cum", "Degas Cum", "Clean Cum", "Runtime Cum"]
        if group == "pause":
            return ["DEGAS_PAUSE", "TREAT_PAUSE", "ERROR", "DEGAS_PAUSE Cum", "TREAT_PAUSE Cum", "ERROR Cum"]
        return list(self.checks.keys())

    def set_names_checked(self, names, checked):
        for n in names:
            cb = self.checks.get(n)
            if cb:
                cb.blockSignals(True)
                cb.setChecked(checked)
                cb.blockSignals(False)

    def sync_group_checks(self):
        groups = [("daily", getattr(self, "daily_group_check", None)),
                  ("cum", getattr(self, "cum_group_check", None)),
                  ("pause", getattr(self, "pause_group_check", None))]
        all_values = []
        for group, parent in groups:
            if not parent:
                continue
            vals = [self.checks[n].isChecked() for n in self.group_names(group) if n in self.checks]
            all_values.extend(vals)
            parent.blockSignals(True)
            parent.setChecked(bool(vals) and all(vals))
            parent.blockSignals(False)
        if hasattr(self, "all_check"):
            self.all_check.blockSignals(True)
            self.all_check.setChecked(bool(all_values) and all(all_values))
            self.all_check.blockSignals(False)

    def set_all(self):
        st = self.all_check.isChecked()
        self.set_names_checked(self.group_names("all"), st)
        for parent in [getattr(self, "daily_group_check", None), getattr(self, "cum_group_check", None), getattr(self, "pause_group_check", None)]:
            if parent:
                parent.blockSignals(True)
                parent.setChecked(st)
                parent.blockSignals(False)
        self.update_chart()

    def set_group(self, group):
        parent = {"daily": getattr(self, "daily_group_check", None),
                  "cum": getattr(self, "cum_group_check", None),
                  "pause": getattr(self, "pause_group_check", None)}.get(group)
        if not parent:
            return
        self.set_names_checked(self.group_names(group), parent.isChecked())
        self.sync_group_checks()
        self.update_chart()

    def on_child_check_changed(self):
        self.sync_group_checks()
        self.update_chart()

    def show_analyze_performance(self):
        msg = getattr(self, "last_performance_report", "")
        if not msg:
            msg = "No performance report yet. Run Analyze first."
        QMessageBox.information(self, APP_NAME, msg)

    def show_analyze_debug(self):
        d = getattr(AnalyzeWorker, "last_debug", {})
        if not d:
            QMessageBox.information(self, APP_NAME, "No analyze debug data yet. Run Analyze / Update first.")
            return
        msg = (
            "Analyze debug\n\n"
            "Scanned: %s\n"
            "Matched: %s\n"
            "Parse failed: %s\n"
            "Duplicates: %s\n\n"
            "First matched files:\n%s\n\n"
            "First errors:\n%s"
        ) % (
            d.get("scanned", ""),
            d.get("matched", ""),
            d.get("parse_fail", ""),
            d.get("duplicate", ""),
            "\n".join(d.get("first_files", [])[:10]) if d.get("first_files") else "(none)",
            "\n".join(d.get("errors", [])[:10]) if d.get("errors") else "(none)",
        )
        QMessageBox.information(self, APP_NAME, msg)

    def date_parse_check(self):
        if not self.folder:
            self.select_folder()
        if not self.folder:
            return
        try:
            files = collect_files(self.folder)
            ok = []
            ng = []
            for p in files[:300]:
                d = build_file_datetime_text(p)
                if d:
                    ok.append("%s -> %s" % (os.path.basename(p), d.strftime("%Y/%m/%d %H:%M:%S")))
                else:
                    ng.append(os.path.basename(p))
            msg = (
                "Date Parse Check\n\n"
                "Candidate files: %d\n"
                "Parsed date: %d\n"
                "No date: %d\n\n"
                "First parsed:\n%s\n\n"
                "First no-date files:\n%s"
            ) % (
                len(files),
                len(ok),
                len(ng),
                "\n".join(ok[:10]) if ok else "(none)",
                "\n".join(ng[:10]) if ng else "(none)",
            )
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            self.hide_wait()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def scan_folder_check(self):
        if not self.folder:
            self.select_folder()
        if not self.folder:
            return
        try:
            d = scan_folder_diagnostics(self.folder)
            msg = "Folder scan check\n\nFolder:\n%s\n\nTotal files: %d\n.txt/.log files: %d\nCandidate log files: %d\nResultSummary files: %d\n\nFirst candidates:\n%s" % (
                self.folder, d["total"], d["txtlog"], d["accepted"], d.get("resultsummary", 0),
                "\n".join(d["samples"]) if d["samples"] else "(none)"
            )
            self.status.setText("Scan: total=%d / txt-log=%d / log candidates=%d / ResultSummary=%d" % (d["total"], d["txtlog"], d["accepted"], d.get("resultsummary", 0)))
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def start_analyze(self):
        if not self.folder:
            self.select_folder()
        if not self.folder:
            return
        if self.worker and self.worker.isRunning():
            return

        self.show_progress_dialog("Updating", "Analyzing WaterSystem Logs", "Preparing analysis...", 0, allow_cancel=True)
        self.analyze_start_time = time.time()
        self.set_busy_ui(True)
        self.progress.setValue(0)
        self.status.setText("Starting analysis...\nResultSummary mode: %s" % resultsummary_mode_text(self.result_file, self.use_existing_result_check.isChecked()))
        try:
            self.chart_view.set_data([], {}, None, None, None, None)
            self.chart_view.redraw()
        except Exception:
            pass
        self.btn_analyze.setEnabled(False)
        self.btn_cancel.setEnabled(True)

        self.worker = AnalyzeWorker(self.folder, self.result_file, self.use_existing_result_check.isChecked())
        self.worker.progress.connect(self.on_progress)
        self.worker.finished_ok.connect(self.on_finished)
        self.worker.failed.connect(self.on_failed)
        self.worker.start()

    def cancel_analyze(self):
        if self.worker:
            self.worker.cancel()
            self.status.setText("Cancel requested... waiting for current file/read step to finish.")
            try:
                self.update_progress_dialog("Cancel requested", "Waiting for the current file/read step to finish safely.", 0, 0, None)
            except Exception:
                pass

    def on_progress(self, c, t, n, a, s, stage, elapsed):
        try:
            detail = "%s\nAdded: %d / Skipped: %d" % (n, a, s)
            self.update_progress_dialog(stage, detail, c, t, elapsed)
        except Exception:
            pass
        if t > 0:
            if self.progress.maximum() == 0:
                self.progress.setRange(0, 100)
            self.progress.setValue(int(c * 100 / t))
        else:
            self.progress.setRange(0, 0)
        self.status.setText("%s\n%d / %d\n%s\nAdded: %d  Skipped: %d  Elapsed: %.1fs" % (stage, c, t, n, a, s, elapsed))

    def on_finished(self, rows, summary, added, skipped):
        self.close_progress_dialog()
        self.hide_wait2()
        self.progress.setRange(0, 100)
        self.progress.setValue(100)

        self.rows = rows
        rebuilt = build_summary_rows(rows)
        self.summary_rows = rebuilt if rebuilt else (summary or [])
        self.refresh_dropdowns()
        self.update_chart()
        self.collapse_menu_after_chart_ready()

        self.set_busy_ui(False)
        self.btn_analyze.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.status.setText("Complete and chart redrawn. Mode: %s / Result rows: %d / Chart rows: %d / Added: %d / Skipped: %d / Scanned: %s / Matched: %s / ParseFail: %s / Duplicate: %s" % (getattr(AnalyzeWorker, "last_debug", {}).get("mode", "?"), len(rows), len(self.summary_rows), added, skipped, getattr(AnalyzeWorker, "last_debug", {}).get("scanned", "?"), getattr(AnalyzeWorker, "last_debug", {}).get("matched", "?"), getattr(AnalyzeWorker, "last_debug", {}).get("parse_fail", "?"), getattr(AnalyzeWorker, "last_debug", {}).get("duplicate", "?")))

        try:
            self.show_performance_report(rows, self.summary_rows, added, skipped, time.time() - getattr(self, "analyze_start_time", time.time()))
        except Exception:
            pass

    def on_failed(self, e):
        self.progress.setRange(0, 100)
        self.btn_analyze.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.status.setText("Error")
        QMessageBox.critical(self, APP_NAME, e)

    def refresh_dropdowns(self):
        vals = ["Auto"]
        for r in self.summary_rows:
            d = r.get("DateTime")
            if isinstance(d, dt.datetime):
                vals.append(d.strftime("%Y/%m/%d %H:%M"))
        vals = list(dict.fromkeys(vals))

        for c in [self.x_start, self.x_end, self.cum_start, self.cum_end]:
            c.blockSignals(True)
            c.clear()
            c.addItems(vals)
            c.setCurrentIndex(0)
            c.blockSignals(False)

        self.display_size.blockSignals(True)
        self.display_size.setCurrentText("Auto Fit")
        self.display_size.blockSignals(False)

    def parse_choice(self, c):
        s = c.currentText().strip()
        if not s or s == "Auto":
            return None
        try:
            return dt.datetime.strptime(s, "%Y/%m/%d %H:%M")
        except Exception:
            return None

    def update_chart(self):
        if not hasattr(self, "chart_view"):
            return

        self.apply_display_size()

        if not self.summary_rows and getattr(self, "rows", None):
            self.summary_rows = build_summary_rows(self.rows)
            if self.summary_rows:
                self.refresh_dropdowns()

        if not self.summary_rows:
            self.chart_view.set_data([], {}, None, None, None, None)
            try:
                self.chart_view.redraw()
            except Exception:
                pass
            self.status.setText(
                "No chart data. Result rows=%d / Chart rows=0. Check Date Parse Check and Show Analyze Debug."
                % len(getattr(self, "rows", []) or [])
            )
            return

        all_x = [r["DateTime"] for r in self.summary_rows if isinstance(r.get("DateTime"), dt.datetime)]
        if not all_x:
            self.chart_view.set_data(self.summary_rows, {}, None, None, None, None)
            self.status.setText("No datetime data for chart.")
            return

        # Step6.3: X range is controlled by direct chart interactions.
        x0 = getattr(self.chart_view, "manual_x0", None) or min(all_x)
        x1 = getattr(self.chart_view, "manual_x1", None) or max(all_x)
        if x1 <= x0:
            x0, x1 = min(all_x), max(all_x)

        c0 = self.parse_choice(self.cum_start)
        c1 = self.parse_choice(self.cum_end)

        check_state = {}
        for name, cb in self.checks.items():
            check_state[name] = cb.isChecked()

        self.chart_view.set_data(self.summary_rows, check_state, x0, x1, c0, c1)
        try:
            self.chart_view.redraw()
        except Exception:
            pass

        visible_count = 0
        for r in self.summary_rows:
            d = r.get("DateTime")
            if isinstance(d, dt.datetime) and x0 <= d <= x1:
                visible_count += 1

        self.status.setText("Chart refreshed. Result rows: %d / Chart rows: %d / Visible rows: %d / X range: %s - %s" % (
            len(getattr(self, "rows", []) or []),
            len(self.summary_rows),
            visible_count,
            x0.strftime("%Y/%m/%d %H:%M") if x0 else "Auto",
            x1.strftime("%Y/%m/%d %H:%M") if x1 else "Auto"
        ))

    def show_wait(self, text="Updating... Please wait."):
        try:
            self._wait = QProgressDialog(text, None, 0, 0, self)
            self._wait.setWindowTitle("Updating")
            self._wait.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            self._wait.setMinimumDuration(0)
            self._wait.setCancelButton(None)
            self._wait.show()
            QApplication.processEvents()
        except Exception:
            self._wait = None

    def hide_wait(self):
        try:
            if getattr(self, "_wait", None):
                self._wait.close()
                self._wait = None
            QApplication.processEvents()
        except Exception:
            pass

    def set_busy_ui(self, busy):
        try:
            for b in self.panel.findChildren(QPushButton):
                if getattr(self, "btn_cancel", None) is b:
                    continue
                b.setEnabled(not busy)
            if hasattr(self, "btn_analyze"):
                self.btn_analyze.setEnabled(not busy)
        except Exception:
            pass

    def show_wait2(self, text="Updating... Please wait."):
        try:
            self._wait2 = QtWidgets.QDialog(self)
            self._wait2.setWindowTitle("Updating")
            self._wait2.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            self._wait2.setMinimumSize(420, 170)
            lay = QVBoxLayout(self._wait2)
            title = QLabel("<b>WaterSystem Analyzer</b>")
            title.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(title)
            self._wait2_label = QLabel(text)
            self._wait2_label.setWordWrap(True)
            self._wait2_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(self._wait2_label)
            bar = QProgressBar()
            bar.setRange(0, 0)
            lay.addWidget(bar)
            self._wait2.show()
            self._wait2.raise_()
            self._wait2.activateWindow()
            for _ in range(8):
                QApplication.processEvents()
        except Exception:
            self._wait2 = None

    def update_wait2(self, text):
        try:
            if getattr(self, "_wait2_label", None):
                self._wait2_label.setText(text)
            QApplication.processEvents()
        except Exception:
            pass

    def hide_wait2(self):
        try:
            if getattr(self, "_wait2", None):
                self._wait2.close()
                self._wait2 = None
            self._wait2_label = None
            QApplication.processEvents()
        except Exception:
            pass

    def project_backup(self):
        try:
            dest = QtWidgets.QFileDialog.getExistingDirectory(self, "Select folder for project backup", self.backup_root_dir())
            if not dest:
                return
            self.show_progress_dialog("Backup", "Creating Project Backup", "Collecting Data and Master files...", 0, allow_cancel=False)
            backup_zip = self.create_project_backup_to(dest, reason="Manual")
            self.close_progress_dialog()
            QMessageBox.information(self, APP_NAME, "Backup created:\n%s" % backup_zip)
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())


    def show_progress_dialog(self, title="Updating", step="Starting...", detail="", total=0, allow_cancel=True):
        try:
            self._prog_total = max(0, int(total or 0))
            self._prog_start_time = time.time()
            self._prog_dialog = QtWidgets.QDialog(self)
            self._prog_dialog.setWindowTitle(title)
            self._prog_dialog.setWindowModality(QtCore.Qt.WindowModality.ApplicationModal)
            self._prog_dialog.setMinimumSize(520, 260)
            lay = QVBoxLayout(self._prog_dialog)

            title_label = QLabel("<b>WaterSystem Analyzer</b>")
            title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(title_label)

            self._prog_step_label = QLabel(step)
            self._prog_step_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self._prog_step_label.setStyleSheet("font-weight: bold;")
            lay.addWidget(self._prog_step_label)

            self._prog_detail_label = QLabel(detail or "")
            self._prog_detail_label.setWordWrap(True)
            self._prog_detail_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(self._prog_detail_label)

            self._prog_count_label = QLabel("")
            self._prog_count_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(self._prog_count_label)

            self._prog_bar = QProgressBar()
            if self._prog_total > 0:
                self._prog_bar.setRange(0, self._prog_total)
                self._prog_bar.setValue(0)
            else:
                self._prog_bar.setRange(0, 0)
            lay.addWidget(self._prog_bar)

            if allow_cancel:
                btns = QHBoxLayout()
                btns.addStretch(1)
                self._prog_cancel_button = QPushButton("Cancel")
                self._prog_cancel_button.clicked.connect(self.cancel_analyze)
                btns.addWidget(self._prog_cancel_button)
                btns.addStretch(1)
                lay.addLayout(btns)

            self._prog_dialog.show()
            self._prog_dialog.raise_()
            self._prog_dialog.activateWindow()
            for _ in range(8):
                QApplication.processEvents()
        except Exception:
            self._prog_dialog = None

    def update_progress_dialog(self, step="", detail="", cur=0, total=0, elapsed=None):
        try:
            if not getattr(self, "_prog_dialog", None):
                return
            if step:
                self._prog_step_label.setText(step)
            if detail:
                self._prog_detail_label.setText(detail)
            total = int(total or getattr(self, "_prog_total", 0) or 0)
            cur = int(cur or 0)
            if total > 0:
                self._prog_bar.setRange(0, total)
                self._prog_bar.setValue(max(0, min(cur, total)))
                pct = int((cur / max(1, total)) * 100)
            else:
                pct = 0
            if elapsed is None:
                elapsed = time.time() - getattr(self, "_prog_start_time", time.time())
            remain = ""
            if total > 0 and cur > 0:
                rate = elapsed / max(1, cur)
                remain_sec = max(0, (total - cur) * rate)
                remain = " / Remaining: %s" % self.format_duration(remain_sec)
            self._prog_count_label.setText("Progress: %d / %d  (%d%%) / Elapsed: %s%s" % (
                cur, total, pct, self.format_duration(elapsed), remain
            ))
            QApplication.processEvents()
        except Exception:
            pass

    def close_progress_dialog(self):
        try:
            if getattr(self, "_prog_dialog", None):
                self._prog_dialog.close()
                self._prog_dialog = None
            QApplication.processEvents()
        except Exception:
            pass

    def format_duration(self, seconds):
        try:
            seconds = int(max(0, seconds))
            h = seconds // 3600
            m = (seconds % 3600) // 60
            s = seconds % 60
            if h:
                return "%d:%02d:%02d" % (h, m, s)
            return "%02d:%02d" % (m, s)
        except Exception:
            return "00:00"

    def show_performance_report(self, rows, summary, added, skipped, elapsed):
        try:
            debug = getattr(AnalyzeWorker, "last_debug", {}) or {}
            msg = (
                "Analyze Completed\n\n"
                "Result rows: %d\n"
                "Chart rows: %d\n"
                "Added: %d\n"
                "Skipped: %d\n"
                "Scanned: %s\n"
                "Matched: %s\n"
                "ParseFail: %s\n"
                "Duplicate: %s\n\n"
                "Total elapsed: %s\n"
                "Mode: %s"
            ) % (
                len(rows), len(summary), added, skipped,
                debug.get("scanned", "?"), debug.get("matched", "?"),
                debug.get("parse_fail", "?"), debug.get("duplicate", "?"),
                self.format_duration(elapsed), debug.get("mode", "?"),
            )
            self.last_performance_report = msg
            self.status.setToolTip(msg)
        except Exception:
            pass

    def show_data_location(self):
        try:
            msg = (
                "Current Data Folder\n\n%s\n\n"
                "Default Data Folder\n\n%s\n\n"
                "Settings File\n\n%s"
            ) % (ensure_data_folders(), default_data_folder(), app_settings_path())
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def move_data_folder(self):
        try:
            current = ensure_data_folders()
            new_root = QtWidgets.QFileDialog.getExistingDirectory(
                self,
                "Select new Data folder location",
                default_data_folder()
            )
            if not new_root:
                return
            new_root = os.path.abspath(new_root)
            if os.path.abspath(current).lower() == new_root.lower():
                QMessageBox.information(self, APP_NAME, "Selected folder is already the current Data folder.")
                return

            ret = QMessageBox.question(
                self,
                APP_NAME,
                "Move Data folder?\n\nCurrent:\n%s\n\nNew:\n%s\n\nData will be copied, verified, and the application setting will be updated." % (current, new_root)
            )
            if ret != QMessageBox.StandardButton.Yes:
                return

            self.show_progress_dialog("Move Data", "Moving Data Folder", "Copying Data folder...", 0, allow_cancel=False)
            os.makedirs(new_root, exist_ok=True)
            shutil.copytree(current, new_root, dirs_exist_ok=True)

            self.update_progress_dialog("Move Data", "Updating DataFolder setting...", 0, 0)
            set_data_root(new_root)
            ensure_data_folders()

            self.close_progress_dialog()
            delete_ret = QMessageBox.question(
                self,
                APP_NAME,
                "Data folder moved successfully.\n\nNew Data folder:\n%s\n\nDelete old Data folder?\n\n%s" % (new_root, current)
            )
            if delete_ret == QMessageBox.StandardButton.Yes:
                try:
                    shutil.rmtree(current)
                except Exception as e:
                    QMessageBox.warning(self, APP_NAME, "Could not delete old Data folder:\n%s" % e)

            QMessageBox.information(self, APP_NAME, "Data folder is now:\n%s\n\nRestart is recommended." % new_root)
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def import_release_package(self):
        try:
            # Default selection starts from the application folder, not the current custom Data folder.
            release_dir = QtWidgets.QFileDialog.getExistingDirectory(
                self,
                "Select WaterSystemAnalyzer Release Package folder",
                app_base_dir()
            )
            if not release_dir:
                return
            release_dir = os.path.abspath(release_dir)
            if not os.path.isdir(os.path.join(release_dir, "Data")):
                QMessageBox.warning(self, APP_NAME, "Invalid Release Package.\n\nData folder was not found:\n%s" % release_dir)
                return

            install_dir = QtWidgets.QFileDialog.getExistingDirectory(
                self,
                "Select install/import destination folder",
                app_base_dir()
            )
            if not install_dir:
                return
            install_dir = os.path.abspath(install_dir)

            running_dir = os.path.abspath(app_base_dir())
            release_abs = os.path.abspath(release_dir)
            if install_dir.lower() == running_dir.lower():
                QMessageBox.warning(self, APP_NAME, "Cannot import directly into the running application folder.\n\nPlease select a new destination folder, for example:\nC:\\WaterSystemAnalyzer_New")
                return
            if install_dir.lower() == release_abs.lower() or release_abs.lower().startswith(install_dir.lower() + os.sep):
                QMessageBox.warning(self, APP_NAME, "Destination cannot be the same as the selected Release Package folder or its parent.\n\nPlease select an empty/new install folder.")
                return

            ret = QMessageBox.question(
                self,
                APP_NAME,
                "Import Release Package?\n\nRelease:\n%s\n\nDestination:\n%s\n\nFiles will be copied to the destination. DataFolder will default to Destination\\Data." % (release_dir, install_dir)
            )
            if ret != QMessageBox.StandardButton.Yes:
                return

            self.show_progress_dialog("Import Release", "Importing Release Package", "Copying release files...", 0, allow_cancel=False)
            os.makedirs(install_dir, exist_ok=True)

            for name in os.listdir(release_dir):
                src = os.path.join(release_dir, name)
                dst = os.path.join(install_dir, name)
                self.update_progress_dialog("Importing Release Package", "Copying: %s" % name, 0, 0)
                if os.path.isdir(src):
                    shutil.copytree(src, dst, dirs_exist_ok=True)
                else:
                    # Do not overwrite a running EXE in the current app folder.
                    try:
                        if os.path.abspath(dst).lower() == os.path.abspath(sys.executable).lower():
                            continue
                    except Exception:
                        pass
                    shutil.copy2(src, dst)

            data_dest = os.path.join(install_dir, "Data")
            self.update_progress_dialog("Importing Release Package", "Setting Data folder...", 0, 0)
            set_data_root(data_dest)
            ensure_data_folders()
            self.close_progress_dialog()

            QMessageBox.information(
                self,
                APP_NAME,
                "Release Package imported.\n\nInstall folder:\n%s\n\nData folder:\n%s\n\nRun Migration Wizard next." % (install_dir, data_dest)
            )
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def backup_root_dir(self):
        d = os.path.join(ensure_data_folders(), "Backups")
        os.makedirs(d, exist_ok=True)
        return d

    def create_project_backup_to(self, dest_folder=None, reason="Manual"):
        if dest_folder is None:
            dest_folder = self.backup_root_dir()
        os.makedirs(dest_folder, exist_ok=True)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_zip = os.path.join(dest_folder, "WaterSystemAnalyzer_Ver42_Backup_%s.zip" % ts)
        data_dir = ensure_data_folders()

        with zipfile.ZipFile(backup_zip, "w", zipfile.ZIP_DEFLATED) as z:
            if os.path.isdir(data_dir):
                for f in Path(data_dir).rglob("*"):
                    if f.is_file():
                        # Avoid nesting backup zips inside backup zips.
                        if "Backups" in f.parts and f.suffix.lower() == ".zip":
                            continue
                        z.write(str(f), "Data/" + str(f.relative_to(data_dir)).replace("\\", "/"))
            info = (
                "WaterSystem Analyzer Ver42 Backup\n"
                "Version: %s\n"
                "Algorithm: %s\n"
                "Created: %s\n"
                "Reason: %s\n"
            ) % (APP_VERSION, ALGORITHM_VERSION, dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S"), reason)
            z.writestr("BACKUP_INFO.txt", info)
        return backup_zip

    def automatic_backup(self, reason="Auto"):
        try:
            return self.create_project_backup_to(self.backup_root_dir(), reason=reason)
        except Exception:
            return ""

    def restore_project_backup(self):
        try:
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self,
                "Select Project Backup ZIP",
                self.backup_root_dir(),
                "Backup ZIP (*.zip);;All Files (*.*)"
            )
            if not path:
                return

            ret = QMessageBox.question(
                self,
                APP_NAME,
                "Restore this backup?\n\n%s\n\nA safety backup will be created first. The Backup folder itself will not be deleted." % path
            )
            if ret != QMessageBox.StandardButton.Yes:
                return

            self.show_progress_dialog("Restore", "Restoring Project Backup", "Creating safety backup...", 0, allow_cancel=False)
            safety = self.create_project_backup_to(self.backup_root_dir(), reason="Before restore")
            data_dir = ensure_data_folders()

            self.update_progress_dialog("Restoring Project Backup", "Extracting backup ZIP...", 0, 0)
            tmp = os.path.join(app_base_dir(), "_restore_tmp")
            if os.path.exists(tmp):
                shutil.rmtree(tmp, ignore_errors=True)
            os.makedirs(tmp, exist_ok=True)

            with zipfile.ZipFile(path, "r") as z:
                z.extractall(tmp)

            backup_data = os.path.join(tmp, "Data")
            if not os.path.isdir(backup_data):
                self.close_progress_dialog()
                QMessageBox.warning(self, APP_NAME, "Invalid backup ZIP. Data folder was not found.")
                shutil.rmtree(tmp, ignore_errors=True)
                return

            self.update_progress_dialog("Restoring Project Backup", "Restoring Data subfolders safely...", 0, 0)

            # Do not delete Data itself and do not delete Data\\Backups/Backup.
            protected = {"backup", "backups"}
            os.makedirs(data_dir, exist_ok=True)

            # Remove/replace only subfolders contained in backup, except backup folders.
            for name in os.listdir(backup_data):
                src = os.path.join(backup_data, name)
                dst = os.path.join(data_dir, name)
                if name.lower() in protected:
                    continue
                if os.path.isdir(src):
                    if os.path.exists(dst):
                        shutil.rmtree(dst, ignore_errors=True)
                    shutil.copytree(src, dst, dirs_exist_ok=True)
                else:
                    try:
                        shutil.copy2(src, dst)
                    except Exception:
                        pass

            shutil.rmtree(tmp, ignore_errors=True)

            # Refresh loaded files/chart where possible.
            self.update_progress_dialog("Restoring Project Backup", "Reloading current data and chart...", 0, 0)
            try:
                ensure_master_foundation_files()
                ensure_master_files(ensure_data_folders())
            except Exception:
                pass
            try:
                if self.result_file and os.path.exists(self.result_file):
                    self.rows, _existing_keys = read_existing_rows(self.result_file)
                    rebuilt_summary = build_summary_rows(self.rows)
                    direct_summary = read_chart_summary_direct(self.result_file)
                    self.summary_rows = rebuilt_summary if rebuilt_summary else (direct_summary or [])
                    self.refresh_dropdowns()
                    self.update_chart()
            except Exception:
                pass

            self.close_progress_dialog()
            QMessageBox.information(
                self,
                APP_NAME,
                "Restore completed.\n\nSafety backup created:\n%s\n\nBackup folders were preserved.\nRestart is recommended." % safety
            )
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())


    def backup_manager(self):
        try:
            bdir = self.backup_root_dir()
            backups = sorted(Path(bdir).glob("WaterSystemAnalyzer_Ver42_Backup_*.zip"), reverse=True)
            if not backups:
                QMessageBox.information(self, APP_NAME, "No backups found.\n\nBackup folder:\n%s" % bdir)
                return

            lines = []
            for i, p in enumerate(backups[:30], 1):
                size_mb = p.stat().st_size / (1024 * 1024)
                lines.append("%02d. %s   %.1f MB" % (i, p.name, size_mb))

            msg = "Backup Manager\n\nBackup folder:\n%s\n\nLatest backups:\n%s\n\nUse Restore Project Backup to restore a selected ZIP." % (bdir, "\n".join(lines))
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def migration_wizard(self):
        try:
            self.show_progress_dialog("Migration Wizard", "Checking migration readiness", "Checking files and permissions...", 0, allow_cancel=False)
            QApplication.processEvents()

            data_dir = ensure_data_folders()
            master_dir = os.path.join(data_dir, "Master")
            required = [
                "Hospital_Master.xlsx",
                "Component_Master.xlsx",
                "SN_Master.xlsx",
                "Hospital_Master.xlsx",
                "Replacement_History.xlsx",
                "Forecast_Master.xlsx",
                "Outlier_Master.xlsx",
            ]

            missing = []
            for fname in required:
                if not os.path.exists(os.path.join(master_dir, fname)):
                    missing.append(fname)

            self.update_progress_dialog("Migration Wizard", "Checking write permission...", 0, 0)
            write_ok = True
            try:
                test_file = os.path.join(data_dir, "_migration_write_test.tmp")
                with open(test_file, "w", encoding="utf-8") as f:
                    f.write("ok")
                os.remove(test_file)
            except Exception:
                write_ok = False

            self.close_progress_dialog()

            if missing or not write_ok:
                QMessageBox.warning(
                    self,
                    APP_NAME,
                    "Migration Check completed with warnings.\n\nMissing:\n%s\n\nWrite permission: %s\n\nRun Validate Masters or Master Foundation to recreate missing files." %
                    ("\n".join(missing) if missing else "(none)", "OK" if write_ok else "NG")
                )
            else:
                QMessageBox.information(
                    self,
                    APP_NAME,
                    "Migration Complete / Ready\n\nData folder:\n%s\n\nMaster files: OK\nWrite permission: OK\n\nYou can start normal operation." % data_dir
                )
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def migration_check(self):
        try:
            checks = []
            base = app_base_dir()
            data_dir = ensure_data_folders()
            master_dir = os.path.join(data_dir, "Master")

            def mark(name, ok, detail=""):
                checks.append((name, ok, detail))

            mark("Application folder", os.path.exists(base), base)
            mark("Data folder", os.path.isdir(data_dir), data_dir)
            mark("Master folder", os.path.isdir(master_dir), master_dir)

            for fname in ["Hospital_Master.xlsx", "Component_Master.xlsx", "SN_Master.xlsx", "Hospital_Master.xlsx", "Replacement_History.xlsx", "Forecast_Master.xlsx", "Outlier_Master.xlsx"]:
                p = os.path.join(master_dir, fname)
                mark(fname, os.path.exists(p), p)

            try:
                test_file = os.path.join(data_dir, "_write_test.tmp")
                with open(test_file, "w", encoding="utf-8") as f:
                    f.write("ok")
                os.remove(test_file)
                mark("Data write permission", True, data_dir)
            except Exception as e:
                mark("Data write permission", False, str(e))

            try:
                import openpyxl
                mark("openpyxl", True, getattr(openpyxl, "__version__", "OK"))
            except Exception as e:
                mark("openpyxl", False, str(e))

            qt_ok = any("Qt" in n or "PySide" in n for n in os.listdir(base)) if os.path.isdir(base) else False
            mark("Qt/PySide files nearby", qt_ok, "Copy the whole dist folder, not only the EXE.")

            ok_count = sum(1 for _n, ok, _d in checks if ok)
            ng = [c for c in checks if not c[1]]
            lines = ["Migration Check", "", "Data folder: %s" % ensure_data_folders(), "Settings: %s" % app_settings_path(), "", "OK: %d / %d" % (ok_count, len(checks)), ""]
            for name, ok, detail in checks:
                lines.append("%s %s" % ("OK " if ok else "NG ", name))
                if detail:
                    lines.append("    %s" % detail)
            if not ng:
                lines.append("")
                lines.append("Ready for another PC. Copy the whole dist folder including Data.")
            else:
                lines.append("")
                lines.append("Warnings found. See NG lines above.")
            QMessageBox.information(self, APP_NAME, "\n".join(lines))
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def export_release_package(self):
        try:
            dest = QtWidgets.QFileDialog.getExistingDirectory(self, "Select output folder for release package", os.getcwd())
            if not dest:
                return
            self.show_wait("Creating release package... Please wait.")
            package = os.path.join(dest, "WaterSystemAnalyzer_Ver42_Release_%s" % dt.datetime.now().strftime("%Y%m%d_%H%M%S"))
            if os.path.exists(package):
                shutil.rmtree(package)
            os.makedirs(package, exist_ok=True)

            base = app_base_dir()
            # Copy app folder contents when running from Nuitka dist; when running source, copy project files.
            for name in os.listdir(base):
                if name in ("build", "build_debug", "__pycache__"):
                    continue
                src = os.path.join(base, name)
                dst = os.path.join(package, name)
                if os.path.isdir(src):
                    if name.lower() == "data":
                        shutil.copytree(src, dst, dirs_exist_ok=True)
                    elif name.endswith(".dist") or name in ("PySide6", "qt-plugins"):
                        shutil.copytree(src, dst, dirs_exist_ok=True)
                else:
                    if name.lower().endswith((".exe", ".dll", ".pyd", ".py", ".bat", ".txt")):
                        shutil.copy2(src, dst)

            data_dir = ensure_data_folders()
            if os.path.isdir(data_dir):
                shutil.copytree(data_dir, os.path.join(package, "Data"), dirs_exist_ok=True)

            with open(os.path.join(package, "WaterSystemAnalyzer_Settings.ini"), "w", encoding="utf-8") as f:
                f.write("DataFolder=%s\n" % os.path.join(package, "Data"))
                f.write("APP_VERSION=%s\n" % APP_VERSION)
            with open(os.path.join(package, "README_RELEASE.txt"), "w", encoding="utf-8") as f:
                f.write("WaterSystem Analyzer Ver42 Release Package\n")
                f.write("Version: %s\n\n" % APP_VERSION)
                f.write("Copy this whole folder to another PC. Do not copy only the EXE.\n")
                f.write("Run Migration Wizard after first launch.\n")
                f.write("Use Restore Project Backup if you need to restore a backup ZIP.\n")
            with open(os.path.join(package, "Version.txt"), "w", encoding="utf-8") as f:
                f.write("APP_VERSION=%s\nALGORITHM_VERSION=%s\nBUILD_DATE=%s\n" % (APP_VERSION, ALGORITHM_VERSION, dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")))
            with open(os.path.join(package, "MigrationGuide.txt"), "w", encoding="utf-8") as f:
                f.write("Migration Guide\n")
                f.write("1. Copy this whole folder to the target PC.\n")
                f.write("2. Do not copy only the EXE.\n")
                f.write("3. Start the EXE.\n")
                f.write("4. Run Migration Wizard.\n")
                f.write("5. Run Migration Check if needed.\n")
                f.write("6. Use Restore Project Backup only when restoring from a backup ZIP.\n")
            with open(os.path.join(package, "Version.txt"), "w", encoding="utf-8") as f:
                f.write("APP_VERSION=%s\nALGORITHM_VERSION=%s\nBUILD_DATE=%s\n" % (APP_VERSION, ALGORITHM_VERSION, dt.datetime.now().strftime("%Y/%m/%d %H:%M:%S")))
            with open(os.path.join(package, "MigrationGuide.txt"), "w", encoding="utf-8") as f:
                f.write("Migration Guide\n")
                f.write("1. Copy this whole folder to the target PC.\n")
                f.write("2. Do not copy only the EXE.\n")
                f.write("3. Start the EXE.\n")
                f.write("4. Run Migration Wizard.\n")
                f.write("5. Run Migration Check if needed.\n")
                f.write("6. Use Restore Project Backup only when restoring from a backup ZIP.\n")

            self.hide_wait()
            QMessageBox.information(self, APP_NAME, "Release package created:\n%s" % package)
        except Exception:
            self.hide_wait()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def dashboard_preview_action(self):
        try:
            if not self.summary_rows and getattr(self, "rows", None):
                self.summary_rows = build_summary_rows(self.rows)
            mode = "Unknown"
            try:
                mode = self.missing_log_handling.currentText()
            except Exception:
                pass
            self.show_progress_dialog("Dashboard Preview", "Building Dashboard Preview", "Updating reliability context...", 0, allow_cancel=False)
            try:
                if getattr(self, "summary_rows", None) and "save_runtime_reconstruction" in globals():
                    save_runtime_reconstruction(self.summary_rows, mode)
            except Exception:
                pass
            latest, trend_rows, explain_rows = dashboard_context(getattr(self, "summary_rows", []))
            self.close_progress_dialog()
            DashboardPreviewDialog(latest, trend_rows, explain_rows, self).exec()
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def data_foundation_action(self):
        try:
            self.show_progress_dialog("Data Foundation", "Initializing Data Foundation", "Creating project metadata and master files...", 0, allow_cancel=False)
            report, paths = build_data_foundation_report()
            self.close_progress_dialog()
            DataFoundationDialog(report, paths, self).exec()
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def open_master_foundation(self):
        try:
            ensure_master_foundation_files()
            dlg = MasterFoundationDialog(self)
            dlg.exec()
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def open_data_folder(self):
        try:
            path = ensure_data_folders()
            os.startfile(path)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def validate_masters(self):
        try:
            ensure_data_folders()
            files = [master_path("Hospital_Master.xlsx"), master_path("Replacement_History.xlsx"), master_path("Forecast_Master.xlsx"), master_path("Outlier_Master.xlsx")]
            msg = "Master validation\n\n"
            for p in files:
                msg += "%s : %s\n" % (os.path.basename(p), "OK" if os.path.exists(p) else "MISSING")
            msg += "\nData folder:\n%s" % ensure_data_folders()
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def runtime_reconstruction_action(self):
        try:
            if not self.summary_rows and getattr(self, "rows", None):
                self.summary_rows = build_summary_rows(self.rows)
            if not self.summary_rows:
                QMessageBox.warning(self, APP_NAME, "No chart/log summary data loaded. Run Analyze or load ResultSummary first.")
                return

            mode = "Unknown"
            try:
                mode = self.missing_log_handling.currentText()
            except Exception:
                pass

            self.show_progress_dialog("Runtime Reconstruction", "Reconstructing Runtime", "Building measured / estimated / unknown runtime basis...", 0, allow_cancel=False)
            paths, runtime_rows, gap_rows, prediction_rows, confidence_rows = save_runtime_reconstruction(self.summary_rows, mode)
            self.close_progress_dialog()

            RuntimeReconstructionDialog(runtime_rows, gap_rows, prediction_rows, confidence_rows, self).exec()
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def reliability_dashboard_action(self):
        try:
            if not self.summary_rows and getattr(self, "rows", None):
                self.summary_rows = build_summary_rows(self.rows)
            if not self.summary_rows:
                QMessageBox.warning(self, APP_NAME, "No chart/log summary data loaded. Run Analyze or load ResultSummary first.")
                return

            mode = "Unknown"
            try:
                mode = self.missing_log_handling.currentText()
            except Exception:
                pass

            self.show_progress_dialog("Reliability", "Building Reliability Dashboard", "Validating replacement history and log gaps...", 0, allow_cancel=False)
            paths, validation_rows, gap_rows, reliability_rows, snapshot_rows = save_reliability_dashboard(self.summary_rows, mode)
            self.close_progress_dialog()

            ReliabilityDashboardDialog(validation_rows, gap_rows, reliability_rows, snapshot_rows, self).exec()
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def import_replacement_history_action(self):
        try:
            # In Ver7.1 this uses the same History.xlsx import engine, but is exposed in Forecast workflow.
            # A full preview/commit wizard will be added after this foundation.
            ensure_master_foundation_files()
            dlg = MasterFoundationDialog(self)
            QMessageBox.information(
                self,
                APP_NAME,
                "Import Replacement History is currently handled in Master Foundation.\n\nClick Import History.xlsx in the dialog.\n\nMapping:\nColumn A = Serial Number\nColumn B = Hospital Name / Site\nColumn C onward = replacement dates"
            )
            dlg.exec()
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def forecast_foundation_action(self):
        try:
            if not self.summary_rows and getattr(self, "rows", None):
                self.summary_rows = build_summary_rows(self.rows)
            if not self.summary_rows:
                QMessageBox.warning(self, APP_NAME, "No chart/log summary data loaded. Run Analyze or load ResultSummary first.")
                return
            mode = self.missing_log_handling.currentText() if hasattr(self, "missing_log_handling") else "Unknown"
            self.show_progress_dialog("Forecast Foundation", "Building Data Coverage", "Checking logs and summary data...", 0, allow_cancel=False)
            coverage_rows, history_rows = save_forecast_foundation(self.summary_rows, mode)
            self.close_progress_dialog()
            ForecastFoundationDialog(coverage_rows, history_rows, self).exec()
        except Exception:
            self.close_progress_dialog()
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def calculate_forecast_now(self):
        if not self.summary_rows:
            QMessageBox.warning(self, APP_NAME, "No chart data. Analyze logs or select a ResultSummary first.")
            return
        default_sn = "NA"
        if self.result_file:
            m = re.search(r"ResultSummary_(.+?)(?:_\d{8}_\d{6})?\.xlsx$", os.path.basename(self.result_file), re.I)
            if m:
                default_sn = m.group(1)
        sn, ok = QInputDialog.getText(self, APP_NAME, "Serial Number:", text=default_sn)
        if not ok:
            return
        try:
            f1 = append_forecast_master(sn or "NA", self.summary_rows, "Failure only", self.forecast_scope.currentText())
            f2 = append_forecast_master(sn or "NA", self.summary_rows, "Preventive included", self.forecast_scope.currentText())
            QMessageBox.information(self, APP_NAME, "Forecast updated.\n\nFailure Only Confidence: %s%%\nAll Replacement Confidence: %s%%" % (f1.get("confidence", ""), f2.get("confidence", "")))
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def resultsummary_check(self):
        try:
            use_existing = self.use_existing_result_check.isChecked() if hasattr(self, "use_existing_result_check") else False
            mode = resultsummary_mode_text(self.result_file, use_existing)
            latest = find_latest_result_summary(self.folder) if self.folder else ""

            resolved = resolve_resultsummary_base_file(self.folder or "", self.result_file, use_existing)
            msg = (
                "ResultSummary check\n\n"
                "Mode: %s\n\n"
                "Selected folder:\n%s\n\n"
                "Explicit selected ResultSummary:\n%s\n\n"
                "Latest ResultSummary in selected folder:\n%s\n\n"
                "Resolved base used by Analyze:\n%s\n\n"
                "Current loaded result rows: %d\n"
                "Current chart rows: %d\n\n"
                "This check does not change Save or chart data."
            ) % (
                mode,
                self.folder or "(none)",
                self.result_file or "(none)",
                latest or "(none)",
                resolved or "(none / log files only)",
                len(self.rows),
                len(self.summary_rows),
            )
            QMessageBox.information(self, APP_NAME, msg)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())

    def show_current_data_count(self):
        dates = [r.get("DateTime") for r in self.summary_rows if isinstance(r.get("DateTime"), dt.datetime)]
        max_daily = 0.0
        for r in self.summary_rows:
            for n in ["Treat", "Degas", "Clean", "Runtime", "DEGAS_PAUSE", "TREAT_PAUSE", "ERROR"]:
                try:
                    max_daily = max(max_daily, float(r.get(n) or 0.0))
                except Exception:
                    pass

        QMessageBox.information(
            self,
            APP_NAME,
            "Current data count\n\n"
            "Result rows: %d\n"
            "Chart rows: %d\n"
            "First chart date: %s\n"
            "Last chart date: %s\n"
            "Max daily value: %.3f\n\n"
            "Last analyze debug:\n%s"
            % (
                len(self.rows),
                len(self.summary_rows),
                min(dates).strftime("%Y/%m/%d %H:%M") if dates else "(none)",
                max(dates).strftime("%Y/%m/%d %H:%M") if dates else "(none)",
                max_daily,
                str(getattr(AnalyzeWorker, "last_debug", {})),
            )
        )

    def clear_change(self):
        for cb in self.checks.values():
            cb.blockSignals(True)
            cb.setChecked(True)
            cb.blockSignals(False)
        self.sync_group_checks()
        for c in [self.x_start, self.x_end, self.cum_start, self.cum_end]:
            c.setCurrentIndex(0)
        self.display_size.setCurrentText("Auto Fit")
        if hasattr(self, "chart_view"):
            self.chart_view.manual_x0 = None
            self.chart_view.manual_x1 = None
        self.update_chart()

    def save_result_summary_action(self):
        if not self.rows:
            QMessageBox.warning(self, APP_NAME, "No data to save.")
            return

        default_sn = "NA"
        if self.result_file:
            m = re.search(r"ResultSummary_(.+?)(?:_\d{8}_\d{6})?\.xlsx$", os.path.basename(self.result_file), re.I)
            if m:
                default_sn = m.group(1)

        sn, ok = QInputDialog.getText(self, APP_NAME, "Serial Number:", text=default_sn)
        if not ok:
            return

        try:
            self.automatic_backup("Before Save ResultSummary")
            self.show_wait("Saving ResultSummary... Please wait.")
            final_path = save_or_update_resultsummary(self.folder or os.getcwd(), self.result_file, self.rows, self.summary_rows, sn or "NA")
            self.result_file = final_path
            self.result_box.setText(final_path)
            self.folder = os.path.dirname(final_path)
            self.folder_box.setText(self.folder)
            self.hide_wait()
            QMessageBox.information(self, APP_NAME, "Updated:\n%s\n\nBackup is created in ResultSummary_Backup when an existing file is overwritten." % final_path)
        except Exception:
            QMessageBox.critical(self, APP_NAME, traceback.format_exc())


def main():
    write_startup_log("Application start")
    try:
        app = QApplication(sys.argv)
        win = AnalyzerWindow()
        win.show()
        return app.exec()
    except Exception:
        try:
            with open(os.path.join(app_base_dir(), "startup_error.log"), "w", encoding="utf-8") as f:
                f.write(traceback.format_exc())
        except Exception:
            pass
        raise


if __name__ == "__main__":
    sys.exit(main())
